from io import BytesIO
from datetime import datetime
import joblib
import pandas as pd
import plotly.express as px
import streamlit as st
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.marker import DataPoint
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.image import Image as XLImage
import matplotlib.pyplot as plt
import re
import math
import json
import hashlib

from pathlib import Path

def validate_student_inputs(name, student_id):
    errors = []

    if student_id.strip():
        if not student_id.isdigit():
            errors.append("❌ Student ID must contain numbers only.")
        elif len(student_id.strip()) != 7:
            errors.append("❌ Student ID must contain exactly 7 digits.")

    return errors

ROOT = Path(__file__).resolve().parent
DATA = ROOT / "dataset" / "Student_data.csv"
MODELS = ROOT / "models"
RESULTS = ROOT / "results"


st.set_page_config(
    page_title="Student Performance Prediction",
    page_icon="🎓",
    layout="wide"
)

st.markdown("""
<style>
:root {
    --sidebar-width: 145px;
    --navy: #0f172a;
    --navy-soft: #172554;
    --blue: #2563eb;
    --violet: #7c3aed;
    --surface: #ffffff;
    --surface-soft: #f8fafc;
    --border: #e2e8f0;
    --text: #172033;
    --muted: #64748b;
}

.stApp {
    background:
        radial-gradient(
            circle at 50% 78%,
            rgba(246, 190, 210, 0.30) 0%,
            rgba(246, 190, 210, 0.14) 22%,
            transparent 42%
        ),
        radial-gradient(
            circle at 18% 14%,
            rgba(132, 171, 235, 0.24) 0%,
            rgba(132, 171, 235, 0.10) 28%,
            transparent 48%
        ),
        linear-gradient(
            180deg,
            #c9d9f3 0%,
            #d7e2f4 26%,
            #dce4f3 50%,
            #d9d8ea 74%,
            #cfc4df 100%
        );
    background-size: cover;
    background-attachment: fixed;
    color: var(--text);
}

@keyframes ambientShift {
    0% {
        background-position: 0% 0%;
    }
    100% {
        background-position: 100% 100%;
    }
}

.block-container {
    max-width: 980px;
    padding-top: .55rem;
    padding-bottom: 2rem;
    background: rgba(255,255,255,0.06);
    border-radius: 28px;
}

/* Sidebar */
[data-testid="stSidebar"] {
    width: 230px !important;
    min-width: 230px !important;

    background:
        radial-gradient(circle at top left, rgba(99,102,241,0.25), transparent 30%),
        linear-gradient(180deg, #0b1220 0%, #111827 55%, #172554 100%);

    border-right: 1px solid rgba(255,255,255,0.08);
}

[data-testid="stSidebar"] > div:first-child {
    width: 230px !important;
    padding: 1.3rem 0.85rem 7rem 0.85rem !important;
}

[data-testid="stSidebar"] * {
    color: white;
}

/* Brand */
[data-testid="stSidebar"] h1 {
    font-size: 1.55rem !important;
    font-weight: 900 !important;
    text-align: center;
    margin-bottom: 1.5rem !important;
}

/* Navigation spacing */
[data-testid="stSidebar"] div[role="radiogroup"] {
    gap: 0.45rem !important;
}

/* Navigation buttons */
[data-testid="stSidebar"] div[role="radiogroup"] label {
    width: 100%;
    min-height: 38px !important;
    display: flex !important;
    align-items: center !important;

    padding: 0.2rem 0.55rem !important;

    border-radius: 12px !important;
    background: rgba(255,255,255,0.06);
    border: 1px solid rgba(255,255,255,0.08);

    transition: .2s ease;
}

[data-testid="stSidebar"] div[role="radiogroup"] label p {
    font-size: 0.78rem !important;
    font-weight: 750 !important;
    white-space: nowrap !important;
    margin: 0 !important;
}

[data-testid="stSidebar"] div[role="radiogroup"] label:hover {
    transform: translateX(4px);
    background: rgba(255,255,255,0.12);
}

[data-testid="stSidebar"] div[role="radiogroup"] label:has(input:checked) {
    background: linear-gradient(135deg,#2563eb,#7c3aed);
    box-shadow: 0 10px 25px rgba(124,58,237,.35);
}

/* Sidebar buttons */
[data-testid="stSidebar"] .stButton button {
    min-height: 34px !important;
    height: 34px !important;
    border-radius: 10px !important;
    font-size: 0.72rem !important;
    font-weight: 750 !important;
}







/* Main hero */
.hero {
    padding: 1.1rem 1.35rem;
    border-radius: 28px;
    color: white;
    background:
        radial-gradient(
            circle at 92% 12%,
            rgba(99, 102, 241, 0.22),
            transparent 28%
        ),
        linear-gradient(
            135deg,
            #0b1220 0%,
            #111827 48%,
            #172554 100%
        );
    margin-bottom: 1.15rem;
    box-shadow:
        0 22px 48px rgba(2, 6, 23, 0.30),
        inset 0 1px 0 rgba(255,255,255,0.06);
    position: relative;
    overflow: hidden;
    border: 1px solid rgba(255,255,255,0.08);
}

.hero:after {
    content: "";
    position: absolute;
    width: 250px;
    height: 250px;
    border-radius: 50%;
    right: -85px;
    top: -105px;
    background: rgba(99,102,241,0.18);
}

.hero h1 {
    margin: 0;
    font-size: 1.75rem;
    font-weight: 850;
    letter-spacing: -0.035em;
    position: relative;
    z-index: 1;
}

.hero p {
    margin-top: .65rem;
    margin-bottom: 0;
    color: #cbd5e1;
    font-size: 0.90rem;
    position: relative;
    z-index: 1;
}

/* Buttons */
.stButton > button,
.stDownloadButton > button,
[data-testid="stFormSubmitButton"] > button {
    width: 100%;
    border: none;
    border-radius: 10px;
    font-weight: 800;
    color: white;
    min-height: 46px;
    background: linear-gradient(90deg, #2563eb, #7c3aed);
    box-shadow: 0 10px 24px rgba(37, 99, 235, 0.18);
    transition: transform .18s ease, box-shadow .18s ease;
}

.stButton > button:hover,
.stDownloadButton > button:hover,
[data-testid="stFormSubmitButton"] > button:hover {
    transform: translateY(-2px);
    box-shadow: 0 14px 30px rgba(37, 99, 235, 0.26);
}

/* High-end black style for Make Another Prediction only */
div[data-testid="stButton"] button[kind="secondary"] {
    background: linear-gradient(135deg, #111827 0%, #020617 100%) !important;
    color: #ffffff !important;
    border: 1px solid #374151 !important;
    box-shadow: 0 10px 24px rgba(2, 6, 23, 0.28) !important;
}

div[data-testid="stButton"] button[kind="secondary"]:hover {
    background: linear-gradient(135deg, #1f2937 0%, #111827 100%) !important;
    border-color: #4b5563 !important;
    transform: translateY(-2px);
    box-shadow: 0 14px 30px rgba(2, 6, 23, 0.38) !important;
}

/* Inputs and cards */
[data-testid="stMetric"] {
    background: rgba(255,255,255,0.92);
    border: 1px solid var(--border);
    border-radius: 10px;
    padding: 0.30rem 0.45rem;
    box-shadow: 0 10px 28px rgba(15,23,42,0.05);
}

[data-testid="stMetricLabel"] p {
    color: var(--muted) !important;
    font-weight: 700 !important;
}

[data-testid="stMetricValue"] {
    color: var(--text) !important;
    font-weight: 800 !important;
}

[data-testid="stMetricValue"] > div {
    font-size: 1.85rem !important;
}

[data-testid="stMetricLabel"] p {
    font-size: 0.88rem !important;
}

h2 {
    font-size: 1.45rem !important;
}

h3 {
    font-size: 1.15rem !important;
}

[data-testid="stForm"] {
    background: rgba(255,255,255,0.92);
    border: 1px solid var(--border);
    border-radius: 22px;
    padding: 0.85rem 1rem 1rem 1rem;
    box-shadow: 0 18px 44px rgba(15,23,42,0.06);
}

[data-testid="stDataFrame"],
[data-testid="stImage"],
[data-testid="stPlotlyChart"] {
    background: white;
    border-radius: 10px;
}

/* CGPA guide */
.cgpa-guide {
    background: rgba(255,255,255,0.95);
    padding: 15px;
    border-radius: 16px;
    border: 1px solid var(--border);
    margin: 14px 0 20px 0;
    box-shadow: 0 14px 34px rgba(15,23,42,0.05);
}

.cgpa-grid {
    display: grid;
    grid-template-columns: repeat(4, 1fr);
    gap: 10px;
}

.cgpa-card {
    padding: 14px;
    border-radius: 10px;
    text-align: center;
    font-size: 0.95rem;
    border: 1px solid rgba(15,23,42,0.04);
}

.cgpa-title {
    font-weight: 850;
    font-size: 0.98rem;
    margin-bottom: 6px;
}

.excellent { background: #dcfce7; }
.good { background: #dbeafe; }
.average { background: #fef3c7; }
.risk { background: #fee2e2; }

/* Headings */
h1, h2, h3 {
    color: var(--text);
    letter-spacing: -0.02em;
}

@media (max-width: 800px) {
    .cgpa-grid {
        grid-template-columns: repeat(2, 1fr);
    }

    .hero h1 {
        font-size: 2rem;
    }
}

/* About page */
.about-hero {
    background: linear-gradient(120deg, #eef2ff, #f5f3ff);
    border: 1px solid #ddd6fe;
    border-radius: 10px;
    padding: 18px 20px;
    margin-bottom: 14px;
    box-shadow: 0 10px 28px rgba(79, 70, 229, 0.08);
}

.about-hero-title {
    font-size: 1.35rem;
    font-weight: 850;
    color: #312e81;
    margin-bottom: 6px;
}

.about-hero-subtitle {
    color: #475569;
    font-size: 0.92rem;
}

.about-card {
    border-radius: 16px;
    padding: 16px;
    min-height: 165px;
    border: 1px solid rgba(148, 163, 184, 0.25);
    box-shadow: 0 10px 24px rgba(15, 23, 42, 0.05);
}

.about-blue {
    background: linear-gradient(135deg, #eff6ff, #dbeafe);
}

.about-green {
    background: linear-gradient(135deg, #f0fdf4, #dcfce7);
}

.about-purple {
    background: linear-gradient(135deg, #faf5ff, #ede9fe);
}

.about-orange {
    background: linear-gradient(135deg, #fff7ed, #ffedd5);
}

.about-yellow {
    background: linear-gradient(135deg, #fffbeb, #fef3c7);
}

.about-red {
    background: linear-gradient(135deg, #fff1f2, #fee2e2);
}

.about-card h4 {
    margin: 0 0 10px 0;
    font-size: 1.02rem;
    color: #172033;
}

.about-list-item {
    background: rgba(255,255,255,0.72);
    border-radius: 10px;
    padding: 8px 10px;
    margin: 7px 0;
    font-size: 0.88rem;
}

.about-model {
    background: rgba(255,255,255,0.70);
    border-radius: 10px;
    padding: 9px 10px;
    margin: 8px 0;
    font-size: 0.88rem;
}

.about-target {
    text-align: center;
    border-radius: 10px;
    padding: 14px 10px;
    min-height: 92px;
    border: 1px solid rgba(148,163,184,0.20);
}

.about-target-title {
    font-size: 1rem;
    font-weight: 850;
    margin-bottom: 5px;
}

.about-target-range {
    font-size: 0.82rem;
    color: #475569;
}

/* Keep overview metric values readable and fully visible */
[data-testid="stMetricValue"] > div {
    font-size: 1.45rem !important;
    line-height: 1.15 !important;
    white-space: normal !important;
}

[data-testid="stMetricLabel"] p {
    font-size: 0.65rem !important;
}

/* Premium Prediction Hub */
.prediction-hub-shell {
    background:
        radial-gradient(circle at top right, rgba(124,58,237,.10), transparent 28%),
        radial-gradient(circle at bottom left, rgba(37,99,235,.08), transparent 26%),
        linear-gradient(180deg, #f8fbff 0%, #f4f7ff 100%);
    border: 1px solid #e2e8f0;
    border-radius: 28px;
    padding: 1.25rem;
    margin-bottom: 1.1rem;
}

.prediction-welcome {
    background:
        linear-gradient(135deg, rgba(15,23,42,.99), rgba(30,58,138,.97), rgba(91,33,182,.94));
    border-radius: 24px;
    padding: 1.05rem 1.35rem;
    margin-bottom: 1.1rem;
    color: white;
    box-shadow: 0 24px 60px rgba(37,99,235,.20);
    position: relative;
    overflow: hidden;
}

.prediction-welcome:after {
    content: "";
    position: absolute;
    width: 220px;
    height: 220px;
    border-radius: 50%;
    right: -80px;
    top: -100px;
    background: rgba(255,255,255,.09);
}

.prediction-eyebrow {
    font-size: .76rem;
    font-weight: 850;
    letter-spacing: .14em;
    text-transform: uppercase;
    color: #c7d2fe;
    margin-bottom: .7rem;
    position: relative;
    z-index: 1;
}

.prediction-welcome h2 {
    color: white !important;
    margin: 0;
    font-size: 1.48rem !important;
    line-height: 1.18;
    position: relative;
    z-index: 1;
}

.prediction-welcome p {
    color: #e2e8f0;
    margin: .5rem 0 0;
    line-height: 1.5;
    font-size: .88rem;
    max-width: 880px;
    position: relative;
    z-index: 1;
}

.prediction-mode-card {
    background: rgba(255,255,255,.98);
    border: 1px solid #e2e8f0;
    border-radius: 22px;
    padding: .82rem .92rem .68rem;
    min-height: 245px;
    height: 245px;
    display: flex;
    flex-direction: column;
    box-shadow: 0 16px 38px rgba(15,23,42,.07);
    margin-bottom: .8rem;
    transition: transform .22s ease, box-shadow .22s ease, border-color .22s ease;
}

.prediction-mode-card:hover {
    transform: translateY(-5px) scale(1.01);
    box-shadow: 0 24px 55px rgba(37,99,235,.16);
    border-color: #c7d2fe;
}

.prediction-mode-icon {
    width: 52px;
    height: 52px;
    border-radius: 15px;
    display: flex;
    align-items: center;
    justify-content: center;
    font-size: 1.62rem;
    margin-bottom: .95rem;
    background: linear-gradient(135deg, #dbeafe, #ede9fe);
    box-shadow: 0 8px 18px rgba(99,102,241,.12);
    transition: transform .22s ease;
}

.prediction-mode-card:hover .prediction-mode-icon {
    transform: translateY(-3px) rotate(-2deg);
}

.prediction-mode-title {
    font-size: 1.18rem;
    font-weight: 850;
    color: #172033;
    margin-bottom: .48rem;
}

.prediction-mode-desc {
    color: #64748b;
    font-size: .85rem;
    line-height: 1.42;
    min-height: 38px;
}

.prediction-mode-feature {
    background: #f8fafc;
    border: 1px solid #eef2f7;
    border-radius: 10px;
    padding: .5rem .65rem;
    margin-top: .48rem;
    color: #334155;
    font-size: .84rem;
    font-weight: 650;
}

.why-system {
    background: rgba(255,255,255,.95);
    border: 1px solid #e2e8f0;
    border-radius: 22px;
    padding: 1.1rem 1.2rem;
    margin-top: 1.15rem;
    box-shadow: 0 12px 30px rgba(15,23,42,.05);
}

.why-system-title {
    font-size: 1.02rem;
    font-weight: 850;
    color: #172033;
    margin-bottom: .8rem;
}

.why-grid {
    display: grid;
    grid-template-columns: repeat(5, 1fr);
    gap: .65rem;
}

.why-item {
    background: linear-gradient(135deg, #eff6ff, #f5f3ff);
    border: 1px solid #dbeafe;
    border-radius: 10px;
    padding: .8rem .7rem;
    text-align: center;
    font-size: .82rem;
    font-weight: 750;
    color: #334155;
}

.system-footer {
    text-align: center;
    color: #64748b;
    font-size: .78rem;
    padding: 1.2rem 0 .2rem;
    line-height: 1.7;
}


.mode-page-hero {
    background:
        radial-gradient(circle at top right, rgba(124,58,237,.10), transparent 30%),
        linear-gradient(120deg, #dcecff 0%, #e7edff 52%, #efe7ff 100%);
    border: 1px solid #dbeafe;
    border-radius: 20px;
    padding: .95rem 1.15rem;
    margin-bottom: .8rem;
    box-shadow: 0 12px 30px rgba(37,99,235,.08);
}

.mode-page-eyebrow {
    font-size: .70rem;
    font-weight: 850;
    letter-spacing: .12em;
    text-transform: uppercase;
    color: #4338ca;
    margin-bottom: .35rem;
}

.mode-page-hero h2 {
    margin: 0;
    color: #172033 !important;
    font-size: 1.28rem !important;
}

.mode-page-hero p {
    margin: .4rem 0 0;
    color: #475569;
    font-size: .86rem;
    line-height: 1.5;
}


.prediction-feature-row {
    display: flex;
    gap: .38rem;
    flex-wrap: wrap;
    margin-top: .5rem;
}
.prediction-feature-chip {
    background: #f8fafc;
    border: 1px solid #e2e8f0;
    border-radius: 999px;
    padding: .28rem .5rem;
    color: #334155;
    font-size: .72rem;
    font-weight: 700;
}
.batch-help-box {
    background: linear-gradient(135deg, #eff6ff, #eef2ff);
    border: 1px solid #dbeafe;
    border-radius: 16px;
    padding: .72rem .85rem;
    color: #334155;
    font-size: .82rem;
    line-height: 1.45;
}

@media (max-width: 900px) {
    .why-grid {
        grid-template-columns: repeat(2, 1fr);
    }
}

/* Premium Glass Surfaces */
[data-testid="stMetric"],
[data-testid="stForm"],
[data-testid="stDataFrame"],
[data-testid="stPlotlyChart"],
.cgpa-guide,
.about-card,
.about-hero,
.batch-card,
.why-system,
.prediction-mode-card {
    background:
        linear-gradient(145deg, rgba(255,255,255,0.88), rgba(244,248,255,0.78)) !important;
    backdrop-filter: blur(18px);
    -webkit-backdrop-filter: blur(18px);
    border: 1px solid rgba(148,163,184,0.24) !important;
    box-shadow:
        0 20px 46px rgba(59,130,246,0.12),
        inset 0 1px 0 rgba(255,255,255,0.90);
}

/* Inputs */
[data-testid="stTextInput"] input,
[data-testid="stNumberInput"] input,
[data-testid="stFileUploader"] section,
[data-baseweb="select"] > div {
    border-radius: 14px !important;
    border: 1px solid #dbe4f0 !important;
    background: rgba(255,255,255,0.90) !important;
    transition: border-color .18s ease, box-shadow .18s ease, transform .18s ease;
}

[data-testid="stTextInput"] input:focus,
[data-testid="stNumberInput"] input:focus,
[data-baseweb="select"] > div:focus-within {
    border-color: #2563eb !important;
    box-shadow: 0 0 0 4px rgba(37,99,235,0.14) !important;
}

/* Primary buttons */
.stButton > button,
[data-testid="stFormSubmitButton"] > button {
    border-radius: 15px !important;
    background: linear-gradient(135deg, #2563eb 0%, #4f46e5 52%, #7c3aed 100%) !important;
    border: 1px solid rgba(255,255,255,0.18) !important;
    box-shadow: 0 14px 28px rgba(79,70,229,0.24) !important;
}

.stButton > button:hover,
[data-testid="stFormSubmitButton"] > button:hover {
    transform: translateY(-3px) scale(1.01) !important;
    box-shadow: 0 18px 34px rgba(79,70,229,0.30) !important;
}

/* Download buttons */
.stDownloadButton > button {
    border-radius: 15px !important;
    background: linear-gradient(135deg, #10b981 0%, #059669 100%) !important;
    color: white !important;
    border: 1px solid rgba(255,255,255,0.18) !important;
    box-shadow: 0 14px 28px rgba(16,185,129,0.22) !important;
}

.stDownloadButton > button:hover {
    transform: translateY(-3px) scale(1.01) !important;
    box-shadow: 0 18px 34px rgba(5,150,105,0.28) !important;
}

/* Sidebar navigation */
[data-testid="stSidebar"] div[role="radiogroup"] label {
    background: rgba(255,255,255,0.05);
    border: 1px solid rgba(255,255,255,0.07);
    box-shadow: inset 0 1px 0 rgba(255,255,255,0.03);
}

[data-testid="stSidebar"] div[role="radiogroup"] label:hover {
    background: rgba(255,255,255,0.10);
    border-color: rgba(255,255,255,0.12);
    transform: translateX(4px);
}

[data-testid="stSidebar"] div[role="radiogroup"] label:has(input:checked) {
    background: linear-gradient(135deg, #2563eb 0%, #7c3aed 100%);
    box-shadow: 0 12px 28px rgba(79,70,229,0.30);
}

/* Premium KPI cards */
[data-testid="stMetric"] {
    border-radius: 20px !important;
    transition: transform .18s ease, box-shadow .18s ease;
}

[data-testid="stMetric"]:hover {
    transform: translateY(-4px);
    box-shadow: 0 22px 48px rgba(37,99,235,0.14) !important;
}

/* Smooth page entrance */
.block-container {
    animation: pageFade .45s ease-out;
}

@keyframes pageFade {
    from {
        opacity: 0;
        transform: translateY(8px);
    }
    to {
        opacity: 1;
        transform: translateY(0);
    }
}

/* Soft scrollbar */
::-webkit-scrollbar {
    width: 10px;
    height: 10px;
}

::-webkit-scrollbar-track {
    background: rgba(226,232,240,0.55);
}

::-webkit-scrollbar-thumb {
    background: linear-gradient(180deg, #94a3b8, #64748b);
    border-radius: 999px;
    border: 2px solid rgba(248,250,252,0.90);
}

/* Reduce motion for accessibility */
@media (prefers-reduced-motion: reduce) {
    .stApp,
    .block-container {
        animation: none !important;
    }

    * {
        transition: none !important;
    }
}

[data-testid="stPlotlyChart"] {
    background:
        linear-gradient(145deg, rgba(247,250,255,0.92), rgba(235,243,255,0.82)) !important;
    border: 1px solid rgba(96,165,250,0.18) !important;
    box-shadow: 0 22px 48px rgba(59,130,246,0.12) !important;
}

[data-testid="stMetric"]:nth-of-type(4n+1) {
    background: linear-gradient(145deg, rgba(239,246,255,0.96), rgba(219,234,254,0.86)) !important;
}
[data-testid="stMetric"]:nth-of-type(4n+2) {
    background: linear-gradient(145deg, rgba(245,243,255,0.96), rgba(237,233,254,0.86)) !important;
}
[data-testid="stMetric"]:nth-of-type(4n+3) {
    background: linear-gradient(145deg, rgba(236,254,255,0.96), rgba(207,250,254,0.82)) !important;
}
[data-testid="stMetric"]:nth-of-type(4n+4) {
    background: linear-gradient(145deg, rgba(240,253,250,0.96), rgba(204,251,241,0.82)) !important;
}


/* FINAL AI DASHBOARD SIDEBAR OVERRIDE */
[data-testid="stSidebar"] {
    width: 230px !important;
    min-width: 230px !important;
}

[data-testid="stSidebar"] > div:first-child {
    width: 230px !important;
    padding: 1.45rem 0.9rem 7rem 0.9rem !important;
}

/* Navigation spacing */
[data-testid="stSidebar"] div[role="radiogroup"] {
    gap: 0.65rem !important;
}

/* Navigation cards */
[data-testid="stSidebar"] div[role="radiogroup"] label {
    min-height: 42px !important;
    padding: 0.35rem 0.55rem !important;
    border-radius: 14px !important;
    display: flex !important;
    align-items: center !important;
}

/* Navigation text */
[data-testid="stSidebar"] div[role="radiogroup"] label p {
    font-size: 0.78rem !important;
    font-weight: 750 !important;
    white-space: nowrap !important;
    margin: 0 !important;
}

/* Selected navigation */
[data-testid="stSidebar"] div[role="radiogroup"] label:has(input:checked) {
    background: linear-gradient(135deg,#2563eb,#7c3aed) !important;
    box-shadow: 0 10px 25px rgba(124,58,237,.35) !important;
}

/* Hover */
[data-testid="stSidebar"] div[role="radiogroup"] label:hover {
    transform: translateX(4px);
    background: rgba(255,255,255,0.12);
}








/* ===== PREMIUM SIDEBAR V2 ===== */

[data-testid="stSidebar"] {
    width: 250px !important;
    min-width: 250px !important;
}

[data-testid="stSidebar"] > div:first-child {
    padding: 1.2rem 0.9rem 6.5rem 0.9rem !important;
}

/* User area */
.sidebar-user-card {
    margin-top: 0.8rem;
    margin-bottom: 1.2rem;
    padding: 10px 14px;
    border-radius: 14px;
    background: rgba(255,255,255,0.08);
    border: 1px solid rgba(255,255,255,0.12);
}

.sidebar-user-name {
    font-size: 0.82rem;
    font-weight: 750;
    color: white;
}

.sidebar-user-role {
    font-size: 0.68rem;
    color: #cbd5e1;
}

/* Navigation */
[data-testid="stSidebar"] div[role="radiogroup"] {
    gap: 0.45rem !important;
}

[data-testid="stSidebar"] div[role="radiogroup"] label {
    min-height: 40px !important;
    padding: 0.25rem 0.55rem !important;
    border-radius: 13px !important;
}

[data-testid="stSidebar"] div[role="radiogroup"] label p {
    font-size: 0.76rem !important;
    font-weight: 750 !important;
}

/* Logout */
[data-testid="stSidebar"] button {
    border-radius: 12px !important;
    min-height: 46px !important;
    font-size: 0.75rem !important;
}






/* ===== FINAL SIDEBAR CLEAN OVERRIDE ===== */

[data-testid="stSidebar"] {
    width: 250px !important;
    min-width: 250px !important;
}

[data-testid="stSidebar"] > div:first-child {
    padding: 1.2rem 0.9rem 4.2rem 0.9rem !important;
}

/* Keep Student AI brand at top */
[data-testid="stSidebar"] h1 {
    margin-top: 0 !important;
    margin-bottom: 1.6rem !important;
    font-size: 1.55rem !important;
}

/* Navigation cleaner */
[data-testid="stSidebar"] div[role="radiogroup"] {
    gap: 0.35rem !important;
}

[data-testid="stSidebar"] div[role="radiogroup"] label {
    min-height: 56px !important;
    height: 56px !important;
    padding: 0.15rem 0.5rem !important;
    border-radius: 12px !important;
}

[data-testid="stSidebar"] div[role="radiogroup"] label p {
    font-size: 0.88rem !important;
    font-weight: 750 !important;
}

/* Logout size */
[data-testid="stSidebar"] .stButton button {
    height: 34px !important;
    min-height: 34px !important;
    font-size: 0.75rem !important;
    border-radius: 12px !important;
}







/* keep footer away from About button */
.sidebar-footer {
    z-index: 0 !important;
}









/* FINAL COMPACT SIDEBAR FOOTER */
.sidebar-footer {
    position: fixed !important;
    bottom: 12px !important;
    left: 12px !important;
    width: 225px !important;
    text-align: center !important;
    white-space: nowrap !important;
    overflow: hidden !important;
    font-size: 0.52rem !important;
    line-height: 1 !important;
    color: rgba(255,255,255,0.75) !important;
    z-index: 5 !important;
}



/* ===== FIXED SIDEBAR NAVIGATION ===== */

[data-testid="stSidebar"] {
    position: fixed !important;
    height: 100vh !important;
    overflow: hidden !important;
}

[data-testid="stSidebar"] > div:first-child {
    height: 100vh !important;
    overflow: hidden !important;
}

/* Keep navigation area stable */
[data-testid="stSidebar"] div[role="radiogroup"] {
    overflow: hidden !important;
}

/* Keep footer fixed */
.sidebar-footer {
    position: fixed !important;
    bottom: 12px !important;
}


.batch-card-added {
    background: rgba(255,255,255,0.96);
    border-radius: 22px;
    padding: 1.2rem;
    box-shadow: 0 16px 38px rgba(15,23,42,.08);
    border: 1px solid #e2e8f0;
}
.batch-card-added h3 {
    color:#172033;
}
.batch-card-added p {
    color:#64748b;
}

/* FINAL SIZE ALIGNMENT ONLY - Individual & Batch Prediction cards */
.prediction-mode-card {
    height: 245px !important;
    min-height: 245px !important;
    display: flex !important;
    flex-direction: column !important;
    justify-content: flex-start !important;
}

.prediction-mode-desc {
    min-height: 52px !important;
}

.prediction-feature-row {
    min-height: 32px !important;
}


/* FINAL BUTTON SIZE BOOST */
.stButton > button,
.stDownloadButton > button,
[data-testid="stFormSubmitButton"] > button {
    min-height: 54px !important;
    font-size: 0.95rem !important;
}

</style>
""", unsafe_allow_html=True)


@st.cache_data(show_spinner=False)
def load_data():
    df = pd.read_csv(DATA)

    df["Performance_Category"] = pd.cut(
        df["Final_CGPA"],
        bins=[-float("inf"), 2.5, 3.0, 3.5, float("inf")],
        labels=["At Risk", "Average", "Good", "Excellent"],
        right=False
    )

    return df


@st.cache_resource(show_spinner=False)
def load_models():
    return {
        "KNN": joblib.load(MODELS / "knn_model.joblib"),
        "SVM": joblib.load(MODELS / "svm_model.joblib"),
        "ANN": joblib.load(MODELS / "ann_model.joblib"),
    }



def make_excel_bytes(dataframe):
    """Create a clean, colour-coded Excel file for batch prediction results."""
    export_df = dataframe.copy()

    # Remove technical confidence/model metadata columns from the downloaded file.
    columns_to_remove = [
        column for column in export_df.columns
        if "Confidence" in column or column == "Best_Model"
    ]
    export_df = export_df.drop(columns=columns_to_remove, errors="ignore")

    preferred_columns = [
        "Student_ID",
        "Student_Name",
        "Number_of_Subjects",
        "Average_Score",
        "Attendance_Pct",
        "Study_Hours_Per_Day",
        "Previous_CGPA",
        "KNN_Prediction",
        "SVM_Prediction",
        "ANN_Prediction",
        "Final_Prediction",
    ]
    ordered_columns = [
        column for column in preferred_columns
        if column in export_df.columns
    ]
    remaining_columns = [
        column for column in export_df.columns
        if column not in ordered_columns
    ]
    export_df = export_df[ordered_columns + remaining_columns]

    friendly_headers = {
        "Student_ID": "Student ID",
        "Student_Name": "Student Name",
        "Number_of_Subjects": "Number of Subjects",
        "Average_Score": "Average Score",
        "Attendance_Pct": "Attendance Rate (%)",
        "Study_Hours_Per_Day": "Study Hours Per Day",
        "Previous_CGPA": "Previous CGPA",
        "KNN_Prediction": "KNN Prediction",
        "SVM_Prediction": "SVM Prediction",
        "ANN_Prediction": "ANN Prediction",
        "Final_Prediction": "FINAL PREDICTION",
    }
    export_df = export_df.rename(columns=friendly_headers)

    output = BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        export_df.to_excel(
            writer,
            index=False,
            sheet_name="Prediction Results",
        )

        workbook = writer.book
        worksheet = writer.sheets["Prediction Results"]

        # Reusable styles.
        navy_fill = PatternFill("solid", fgColor="1E3A8A")
        final_header_fill = PatternFill("solid", fgColor="5B21B6")
        white_font = Font(color="FFFFFF", bold=True)
        body_font = Font(color="172033")
        thin_grey = Side(style="thin", color="D7DEE8")
        cell_border = Border(
            left=thin_grey,
            right=thin_grey,
            top=thin_grey,
            bottom=thin_grey,
        )

        category_fills = {
            "Excellent": PatternFill("solid", fgColor="DCFCE7"),
            "Good": PatternFill("solid", fgColor="DBEAFE"),
            "Average": PatternFill("solid", fgColor="FEF3C7"),
            "At Risk": PatternFill("solid", fgColor="FEE2E2"),
        }
        category_fonts = {
            "Excellent": Font(color="166534", bold=True),
            "Good": Font(color="1D4ED8", bold=True),
            "Average": Font(color="92400E", bold=True),
            "At Risk": Font(color="B91C1C", bold=True),
        }

        # Header styling.
        for cell in worksheet[1]:
            cell.fill = (
                final_header_fill
                if cell.value == "FINAL PREDICTION"
                else navy_fill
            )
            cell.font = white_font
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )
            cell.border = cell_border

        worksheet.row_dimensions[1].height = 30
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions
        worksheet.sheet_view.showGridLines = False

        # Identify key prediction columns.
        header_lookup = {
            cell.value: cell.column
            for cell in worksheet[1]
        }
        final_prediction_col = header_lookup.get("FINAL PREDICTION")
        prediction_headers = {
            "KNN Prediction",
            "SVM Prediction",
            "ANN Prediction",
            "FINAL PREDICTION",
        }

        # Body styling with alternating rows.
        for row_index in range(2, worksheet.max_row + 1):
            alternate_fill = (
                PatternFill("solid", fgColor="F8FAFC")
                if row_index % 2 == 0
                else PatternFill("solid", fgColor="FFFFFF")
            )

            for cell in worksheet[row_index]:
                cell.fill = alternate_fill
                cell.font = body_font
                cell.border = cell_border
                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                )

                column_header = worksheet.cell(
                    row=1,
                    column=cell.column,
                ).value

                if column_header == "Student Name":
                    cell.alignment = Alignment(
                        horizontal="left",
                        vertical="center",
                    )

                if column_header in prediction_headers:
                    category = str(cell.value)
                    if category in category_fills:
                        cell.fill = category_fills[category]
                        cell.font = category_fonts[category]

            # Make Final Prediction the visual focus.
            if final_prediction_col is not None:
                final_cell = worksheet.cell(
                    row=row_index,
                    column=final_prediction_col,
                )
                final_cell.border = Border(
                    left=Side(style="medium", color="7C3AED"),
                    right=Side(style="medium", color="7C3AED"),
                    top=thin_grey,
                    bottom=thin_grey,
                )

        # Friendly numeric formats.
        for header, number_format in {
            "Average Score": "0.00",
            "Attendance Rate (%)": "0.0",
            "Study Hours Per Day": "0.0",
            "Previous CGPA": "0.00",
        }.items():
            column_index = header_lookup.get(header)
            if column_index:
                for row_index in range(2, worksheet.max_row + 1):
                    worksheet.cell(
                        row=row_index,
                        column=column_index,
                    ).number_format = number_format

        # Sensible widths.
        width_map = {
            "Student ID": 14,
            "Student Name": 22,
            "Number of Subjects": 18,
            "Average Score": 15,
            "Attendance Rate (%)": 19,
            "Study Hours Per Day": 21,
            "Previous CGPA": 15,
            "KNN Prediction": 17,
            "SVM Prediction": 17,
            "ANN Prediction": 17,
            "FINAL PREDICTION": 21,
        }

        for column_index in range(1, worksheet.max_column + 1):
            header = worksheet.cell(row=1, column=column_index).value
            worksheet.column_dimensions[
                get_column_letter(column_index)
            ].width = width_map.get(header, 16)

        # Add a compact summary sheet only for completed prediction output.
        if "FINAL PREDICTION" in export_df.columns:
            summary_sheet = workbook.create_sheet("Summary")
            summary_sheet.sheet_view.showGridLines = False

            summary_sheet.merge_cells("A1:D1")
            summary_sheet["A1"] = "Batch Prediction Summary"
            summary_sheet["A1"].fill = navy_fill
            summary_sheet["A1"].font = Font(
                color="FFFFFF",
                bold=True,
                size=16,
            )
            summary_sheet["A1"].alignment = Alignment(
                horizontal="center",
                vertical="center",
            )
            summary_sheet.row_dimensions[1].height = 32

            summary_sheet["A3"] = "Performance Category"
            summary_sheet["B3"] = "Number of Students"
            summary_sheet["C3"] = "Percentage"
            summary_sheet["D3"] = "Recommended Action"

            for cell in summary_sheet[3]:
                cell.fill = final_header_fill
                cell.font = white_font
                cell.alignment = Alignment(horizontal="center")
                cell.border = cell_border

            category_order = [
                "Excellent",
                "Good",
                "Average",
                "At Risk",
            ]
            counts = (
                export_df["FINAL PREDICTION"]
                .value_counts()
                .reindex(category_order, fill_value=0)
            )
            total_students = len(export_df)

            recommendations = {
                "Excellent": "Maintain strong performance",
                "Good": "Continue current progress",
                "Average": "Provide academic guidance",
                "At Risk": "Early intervention required",
            }

            for offset, category in enumerate(category_order, start=4):
                count = int(counts[category])
                percentage = (
                    count / total_students
                    if total_students
                    else 0
                )

                summary_sheet.cell(offset, 1, category)
                summary_sheet.cell(offset, 2, count)
                summary_sheet.cell(offset, 3, percentage)
                summary_sheet.cell(
                    offset,
                    4,
                    recommendations[category],
                )

                for column_index in range(1, 5):
                    cell = summary_sheet.cell(
                        offset,
                        column_index,
                    )
                    cell.border = cell_border
                    cell.alignment = Alignment(
                        horizontal=(
                            "left"
                            if column_index == 4
                            else "center"
                        ),
                        vertical="center",
                    )

                summary_sheet.cell(offset, 1).fill = category_fills[category]
                summary_sheet.cell(offset, 1).font = category_fonts[category]
                summary_sheet.cell(offset, 3).number_format = "0.0%"

            summary_sheet["A9"] = "Total Students"
            summary_sheet["B9"] = total_students
            summary_sheet["A9"].font = Font(bold=True)
            summary_sheet["B9"].font = Font(bold=True)

            summary_sheet.column_dimensions["A"].width = 24
            summary_sheet.column_dimensions["B"].width = 20
            summary_sheet.column_dimensions["C"].width = 16
            summary_sheet.column_dimensions["D"].width = 32
            summary_sheet.freeze_panes = "A3"

    output.seek(0)
    return output.getvalue()


def make_batch_excel_bytes(dataframe):
    """
    Create a polished batch Excel export.

    Sheet 1: Prediction Results
        - Student information and model input values
        - Final Prediction only, clearly highlighted

    Sheet 2: Model Comparison
        - KNN, SVM and ANN predictions for detailed checking

    Sheet 3: Executive Dashboard
        - KPI cards, native Excel pie chart and bar chart

    Sheet 4: Summary
        - Category totals, percentages and recommended actions
    """
    export_df = dataframe.copy()

    # Remove technical columns that are not useful to normal users.
    columns_to_remove = [
        column for column in export_df.columns
        if "Confidence" in column or column == "Best_Model"
    ]
    export_df = export_df.drop(columns=columns_to_remove, errors="ignore")

    main_columns = [
        column for column in [
            "Student_ID",
            "Student_Name",
            "Average_Score",
            "Attendance_Pct",
            "Study_Hours_Per_Day",
            "Previous_CGPA",
            "Final_Prediction",
        ]
        if column in export_df.columns
    ]

    comparison_columns = [
        column for column in [
            "Student_ID",
            "Student_Name",
            "KNN_Prediction",
            "SVM_Prediction",
            "ANN_Prediction",
        ]
        if column in export_df.columns
    ]

    friendly_headers = {
        "Student_ID": "Student ID",
        "Student_Name": "Student Name",
        "Number_of_Subjects": "Number of Subjects",
        "Average_Score": "Average Score",
        "Attendance_Pct": "Attendance Rate (%)",
        "Study_Hours_Per_Day": "Study Hours Per Day",
        "Previous_CGPA": "Previous CGPA",
        "KNN_Prediction": "KNN Prediction",
        "SVM_Prediction": "SVM Prediction",
        "ANN_Prediction": "ANN Prediction",
        "Final_Prediction": "FINAL PREDICTION",
    }

    main_df = export_df[main_columns].rename(columns=friendly_headers)
    comparison_df = export_df[comparison_columns].rename(columns=friendly_headers)

    output = BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        main_df.to_excel(
            writer,
            index=False,
            sheet_name="Prediction Results",
        )
        comparison_df.to_excel(
            writer,
            index=False,
            sheet_name="Model Comparison",
        )

        workbook = writer.book

        # Shared styles.
        navy_fill = PatternFill("solid", fgColor="1E3A8A")
        purple_fill = PatternFill("solid", fgColor="5B21B6")
        light_purple_fill = PatternFill("solid", fgColor="F5F3FF")
        white_fill = PatternFill("solid", fgColor="FFFFFF")
        alternate_fill = PatternFill("solid", fgColor="F8FAFC")
        white_font = Font(color="FFFFFF", bold=True)
        body_font = Font(color="172033")
        subtitle_font = Font(color="64748B", italic=True)
        thin_grey = Side(style="thin", color="D7DEE8")
        purple_side = Side(style="medium", color="7C3AED")

        category_fills = {
            "Excellent": PatternFill("solid", fgColor="DCFCE7"),
            "Good": PatternFill("solid", fgColor="DBEAFE"),
            "Average": PatternFill("solid", fgColor="FEF3C7"),
            "At Risk": PatternFill("solid", fgColor="FEE2E2"),
        }
        category_fonts = {
            "Excellent": Font(color="166534", bold=True),
            "Good": Font(color="1D4ED8", bold=True),
            "Average": Font(color="92400E", bold=True),
            "At Risk": Font(color="B91C1C", bold=True),
        }

        def apply_common_sheet_style(
            worksheet,
            widths,
            highlighted_header=None,
            prediction_headers=None,
        ):
            worksheet.sheet_view.showGridLines = False
            worksheet.freeze_panes = "A2"
            worksheet.auto_filter.ref = worksheet.dimensions
            worksheet.row_dimensions[1].height = 32

            header_lookup = {
                cell.value: cell.column
                for cell in worksheet[1]
            }

            for cell in worksheet[1]:
                cell.fill = (
                    purple_fill
                    if cell.value == highlighted_header
                    else navy_fill
                )
                cell.font = white_font
                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=True,
                )
                cell.border = Border(
                    left=thin_grey,
                    right=thin_grey,
                    top=thin_grey,
                    bottom=thin_grey,
                )

            for row_index in range(2, worksheet.max_row + 1):
                row_fill = (
                    alternate_fill
                    if row_index % 2 == 0
                    else white_fill
                )

                for cell in worksheet[row_index]:
                    cell.fill = row_fill
                    cell.font = body_font
                    cell.alignment = Alignment(
                        horizontal="center",
                        vertical="center",
                    )
                    cell.border = Border(
                        left=thin_grey,
                        right=thin_grey,
                        top=thin_grey,
                        bottom=thin_grey,
                    )

                    column_header = worksheet.cell(
                        row=1,
                        column=cell.column,
                    ).value

                    if column_header == "Student Name":
                        cell.alignment = Alignment(
                            horizontal="left",
                            vertical="center",
                        )

                    if (
                        prediction_headers
                        and column_header in prediction_headers
                    ):
                        category = str(cell.value)
                        if category in category_fills:
                            cell.fill = category_fills[category]
                            cell.font = category_fonts[category]

                if highlighted_header in header_lookup:
                    highlighted_cell = worksheet.cell(
                        row=row_index,
                        column=header_lookup[highlighted_header],
                    )
                    highlighted_cell.border = Border(
                        left=purple_side,
                        right=purple_side,
                        top=thin_grey,
                        bottom=thin_grey,
                    )

            number_formats = {
                "Average Score": "0.00",
                "Attendance Rate (%)": "0.0",
                "Study Hours Per Day": "0.0",
                "Previous CGPA": "0.00",
            }

            for header, number_format in number_formats.items():
                column_index = header_lookup.get(header)
                if column_index:
                    for row_index in range(2, worksheet.max_row + 1):
                        worksheet.cell(
                            row=row_index,
                            column=column_index,
                        ).number_format = number_format

            for column_index in range(1, worksheet.max_column + 1):
                header = worksheet.cell(
                    row=1,
                    column=column_index,
                ).value
                worksheet.column_dimensions[
                    get_column_letter(column_index)
                ].width = widths.get(header, 16)

        # Sheet 1: clean final result.
        results_sheet = writer.sheets["Prediction Results"]
        apply_common_sheet_style(
            results_sheet,
            widths={
                "Student ID": 14,
                "Student Name": 22,
                "Number of Subjects": 18,
                "Average Score": 15,
                "Attendance Rate (%)": 19,
                "Study Hours Per Day": 21,
                "Previous CGPA": 15,
                "FINAL PREDICTION": 22,
            },
            highlighted_header="FINAL PREDICTION",
            prediction_headers={"FINAL PREDICTION"},
        )

        # Sheet 2: detailed algorithm comparison.
        comparison_sheet = writer.sheets["Model Comparison"]
        apply_common_sheet_style(
            comparison_sheet,
            widths={
                "Student ID": 14,
                "Student Name": 22,
                "KNN Prediction": 18,
                "SVM Prediction": 18,
                "ANN Prediction": 18,
            },
            prediction_headers={
                "KNN Prediction",
                "SVM Prediction",
                "ANN Prediction",
            },
        )

        # Add a professional title above the comparison table.
        comparison_sheet.insert_rows(1, amount=3)
        comparison_sheet.merge_cells("A1:E1")
        comparison_sheet["A1"] = "Machine Learning Model Comparison"
        comparison_sheet["A1"].fill = navy_fill
        comparison_sheet["A1"].font = Font(
            color="FFFFFF",
            bold=True,
            size=16,
        )
        comparison_sheet["A1"].alignment = Alignment(
            horizontal="center",
            vertical="center",
        )
        comparison_sheet.row_dimensions[1].height = 30

        comparison_sheet.merge_cells("A2:E2")
        comparison_sheet["A2"] = (
            "Detailed KNN, SVM and ANN predictions for academic review"
        )
        comparison_sheet["A2"].fill = light_purple_fill
        comparison_sheet["A2"].font = subtitle_font
        comparison_sheet["A2"].alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

        # Reapply freeze/filter after inserted title rows.
        comparison_sheet.freeze_panes = "A5"
        comparison_sheet.auto_filter.ref = (
            f"A4:E{comparison_sheet.max_row}"
        )

        # Sheet 3: executive dashboard with native Excel charts.
        dashboard_sheet = workbook.create_sheet(
            "Executive Dashboard",
            1,
        )
        dashboard_sheet.sheet_view.showGridLines = False
        dashboard_sheet.sheet_view.zoomScale = 80

        # Colour palette.
        dashboard_blue = "1E3A8A"
        dashboard_purple = "6D28D9"
        dashboard_light_blue = "EFF6FF"
        dashboard_border = "CBD5E1"

        card_styles = {
            "Total Students": {
                "fill": "E8EEFF",
                "font": "1E3A8A",
            },
            "Excellent": {
                "fill": "DCFCE7",
                "font": "166534",
            },
            "Good": {
                "fill": "DBEAFE",
                "font": "1D4ED8",
            },
            "Average": {
                "fill": "FEF3C7",
                "font": "B45309",
            },
            "At Risk": {
                "fill": "FEE2E2",
                "font": "B91C1C",
            },
        }

        # Fixed column widths provide a predictable layout in Excel and WPS.
        dashboard_widths = {
            "A": 13,
            "B": 13,
            "C": 13,
            "D": 13,
            "E": 13,
            "F": 13,
            "G": 13,
            "H": 13,
            "I": 13,
            "J": 13,
            "K": 13,
            "L": 13,
            "M": 13,
            "N": 13,
            "O": 13,
        }
        for column_letter, width in dashboard_widths.items():
            dashboard_sheet.column_dimensions[column_letter].width = width

        # Report title.
        dashboard_sheet.merge_cells("A1:O2")
        dashboard_sheet["A1"] = "Student Performance Prediction Report"
        dashboard_sheet["A1"].fill = PatternFill(
            "solid",
            fgColor=dashboard_blue,
        )
        dashboard_sheet["A1"].font = Font(
            color="FFFFFF",
            bold=True,
            size=20,
        )
        dashboard_sheet["A1"].alignment = Alignment(
            horizontal="center",
            vertical="center",
        )
        dashboard_sheet["A1"].border = Border(
            bottom=Side(style="medium", color=dashboard_purple),
        )

        dashboard_sheet.merge_cells("A3:L3")
        dashboard_sheet["A3"] = (
            "Batch prediction dashboard generated by the Student AI system"
        )
        dashboard_sheet["A3"].fill = PatternFill(
            "solid",
            fgColor="F3F0FF",
        )
        dashboard_sheet["A3"].font = Font(
            color="5B21B6",
            bold=True,
            italic=True,
            size=10,
        )
        dashboard_sheet["A3"].alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

        dashboard_sheet.merge_cells("M3:O3")
        dashboard_sheet["M3"] = (
            "Generated: " + datetime.now().strftime("%d %b %Y")
        )
        dashboard_sheet["M3"].fill = PatternFill(
            "solid",
            fgColor="E8EEFF",
        )
        dashboard_sheet["M3"].font = Font(
            color=dashboard_blue,
            bold=True,
            size=9,
        )
        dashboard_sheet["M3"].alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

        # Prediction counts.
        total_students = len(export_df)
        category_order = [
            "Excellent",
            "Good",
            "Average",
            "At Risk",
        ]
        counts = (
            export_df["Final_Prediction"]
            .value_counts()
            .reindex(category_order, fill_value=0)
        )

        # Five aligned KPI cards.
        # Rendered as one image to guarantee identical display in Excel and WPS.
        kpi_buffer = BytesIO()

        kpi_labels = [
            "TOTAL STUDENTS",
            "EXCELLENT",
            "GOOD",
            "AVERAGE",
            "AT RISK",
        ]
        kpi_values = [
            total_students,
            int(counts["Excellent"]),
            int(counts["Good"]),
            int(counts["Average"]),
            int(counts["At Risk"]),
        ]
        kpi_subtitles = [
            "Students analysed",
            "High performance",
            "Positive progress",
            "Academic guidance",
            "Early intervention",
        ]
        kpi_backgrounds = [
            "#E8EEFF",
            "#DCFCE7",
            "#DBEAFE",
            "#FEF3C7",
            "#FEE2E2",
        ]
        kpi_text_colours = [
            "#1E3A8A",
            "#166534",
            "#1D4ED8",
            "#B45309",
            "#B91C1C",
        ]

        fig, axes = plt.subplots(
            1,
            5,
            figsize=(15.4, 2.05),
        )
        fig.patch.set_facecolor("white")

        for axis, label, value, subtitle, background, text_colour in zip(
            axes,
            kpi_labels,
            kpi_values,
            kpi_subtitles,
            kpi_backgrounds,
            kpi_text_colours,
        ):
            axis.set_facecolor(background)
            axis.set_xticks([])
            axis.set_yticks([])

            for spine in axis.spines.values():
                spine.set_visible(True)
                spine.set_color("#CBD5E1")
                spine.set_linewidth(1.1)

            axis.text(
                0.5,
                0.76,
                label,
                ha="center",
                va="center",
                fontsize=10,
                fontweight="bold",
                color=text_colour,
                transform=axis.transAxes,
            )
            axis.text(
                0.5,
                0.48,
                f"{value:,}",
                ha="center",
                va="center",
                fontsize=20,
                fontweight="bold",
                color=text_colour,
                transform=axis.transAxes,
            )
            axis.text(
                0.5,
                0.20,
                subtitle,
                ha="center",
                va="center",
                fontsize=8.5,
                fontweight="bold",
                color=text_colour,
                transform=axis.transAxes,
            )

        plt.subplots_adjust(
            left=0.01,
            right=0.99,
            top=0.95,
            bottom=0.05,
            wspace=0.045,
        )
        fig.savefig(
            kpi_buffer,
            format="png",
            dpi=170,
            bbox_inches="tight",
            facecolor="white",
        )
        plt.close(fig)
        kpi_buffer.seek(0)

        kpi_image = XLImage(kpi_buffer)
        kpi_image.width = 1370
        kpi_image.height = 165
        dashboard_sheet.add_image(
            kpi_image,
            "A5",
        )

        # Reserve vertical space for the KPI image.
        for row_number, height in {
            5: 30,
            6: 30,
            7: 30,
            8: 30,
        }.items():
            dashboard_sheet.row_dimensions[row_number].hidden = False
            dashboard_sheet.row_dimensions[row_number].height = height

        # Performance summary table.
        dashboard_sheet.row_dimensions[9].height = 12
        dashboard_sheet.merge_cells("A10:E10")
        dashboard_sheet["A10"] = "Performance Summary"
        dashboard_sheet["A10"].fill = PatternFill(
            "solid",
            fgColor=dashboard_blue,
        )
        dashboard_sheet["A10"].font = Font(
            color="FFFFFF",
            bold=True,
            size=13,
        )
        dashboard_sheet["A10"].alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

        summary_headers = [
            "Performance Category",
            "Students",
            "Percentage",
        ]
        summary_columns = [1, 4, 5]

        dashboard_sheet.merge_cells("A11:C11")
        dashboard_sheet["A11"] = summary_headers[0]
        dashboard_sheet["D11"] = summary_headers[1]
        dashboard_sheet["E11"] = summary_headers[2]

        for cell_reference in ["A11", "D11", "E11"]:
            cell = dashboard_sheet[cell_reference]
            cell.fill = PatternFill(
                "solid",
                fgColor=dashboard_purple,
            )
            cell.font = Font(
                color="FFFFFF",
                bold=True,
                size=10,
            )
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
            )
            cell.border = Border(
                left=Side(style="thin", color=dashboard_border),
                right=Side(style="thin", color=dashboard_border),
                top=Side(style="thin", color=dashboard_border),
                bottom=Side(style="thin", color=dashboard_border),
            )

        for row_index, category in enumerate(category_order, start=12):
            count = int(counts[category])
            percentage = (
                count / total_students
                if total_students
                else 0
            )

            dashboard_sheet.merge_cells(
                start_row=row_index,
                start_column=1,
                end_row=row_index,
                end_column=3,
            )
            dashboard_sheet.cell(
                row=row_index,
                column=1,
                value=category,
            )
            dashboard_sheet.cell(
                row=row_index,
                column=4,
                value=count,
            )
            dashboard_sheet.cell(
                row=row_index,
                column=5,
                value=percentage,
            )

            category_fill = PatternFill(
                "solid",
                fgColor=card_styles[category]["fill"],
            )
            category_font = card_styles[category]["font"]

            for column_index in range(1, 6):
                cell = dashboard_sheet.cell(
                    row=row_index,
                    column=column_index,
                )
                cell.fill = category_fill
                cell.border = Border(
                    left=Side(style="thin", color=dashboard_border),
                    right=Side(style="thin", color=dashboard_border),
                    top=Side(style="thin", color=dashboard_border),
                    bottom=Side(style="thin", color=dashboard_border),
                )
                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                )
                cell.font = Font(
                    color=category_font,
                    bold=(column_index == 1),
                    size=10,
                )

            dashboard_sheet.cell(
                row=row_index,
                column=5,
            ).number_format = "0.0%"

        # Total row.
        dashboard_sheet.merge_cells("A16:C16")
        dashboard_sheet["A16"] = "TOTAL"
        dashboard_sheet["D16"] = total_students
        dashboard_sheet["E16"] = 1

        for cell_reference in ["A16", "D16", "E16"]:
            cell = dashboard_sheet[cell_reference]
            cell.fill = PatternFill(
                "solid",
                fgColor="E0E7FF",
            )
            cell.font = Font(
                color=dashboard_blue,
                bold=True,
                size=10,
            )
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
            )
            cell.border = Border(
                left=Side(style="thin", color=dashboard_border),
                right=Side(style="thin", color=dashboard_border),
                top=Side(style="thin", color=dashboard_border),
                bottom=Side(style="thin", color=dashboard_border),
            )
        dashboard_sheet["E16"].number_format = "0%"

        # Charts use the visible Performance Summary table.
        # This is more compatible with WPS than referencing hidden columns.

        # Pie chart in the centre.
        # A rendered image is embedded instead of a native Excel chart because
        # WPS may add unwanted "Series1" labels to native chart data labels.
        pie_buffer = BytesIO()
        pie_colours = ["#22C55E", "#3B82F6", "#F59E0B", "#EF4444"]

        fig, ax = plt.subplots(
            figsize=(6.2, 4.3),
            facecolor="white",
        )
        fig.patch.set_facecolor("white")
        fig.patch.set_edgecolor("#C9D3E6")
        fig.patch.set_linewidth(2.2)
        wedges, texts, autotexts = ax.pie(
            [int(counts[category]) for category in category_order],
            labels=None,
            autopct="%1.0f%%",
            startangle=90,
            colors=pie_colours,
            pctdistance=0.70,
            wedgeprops={
                "linewidth": 1.2,
                "edgecolor": "white",
            },
        )
        ax.set_title(
            "Prediction Distribution",
            fontsize=14,
            fontweight="bold",
            pad=14,
        )
        ax.legend(
            wedges,
            [
                f"{category} ({int(counts[category])})"
                for category in category_order
            ],
            loc="center left",
            bbox_to_anchor=(1.00, 0.5),
            frameon=False,
            fontsize=9,
        )
        for autotext in autotexts:
            autotext.set_fontsize(9)
            autotext.set_fontweight("bold")
            autotext.set_color("white")
        ax.axis("equal")
        ax.set_facecolor("white")
        for spine in ax.spines.values():
            spine.set_visible(True)
            spine.set_color("#D6DDEB")
            spine.set_linewidth(1.1)
        plt.tight_layout(pad=0.8)
        fig.savefig(
            pie_buffer,
            format="png",
            dpi=190,
            bbox_inches="tight",
            facecolor="white",
            edgecolor="#C9D3E6",
        )
        plt.close(fig)
        pie_buffer.seek(0)

        pie_image = XLImage(pie_buffer)
        pie_image.width = 420
        pie_image.height = 260
        dashboard_sheet.add_image(
            pie_image,
            "G10",
        )

        # Recommended-action panel on the right.
        dashboard_sheet.merge_cells("J10:N17")
        at_risk_count = int(counts["At Risk"])

        if at_risk_count > 0:
            recommendation_text = (
                "RECOMMENDED ACTION\n\n"
                f"{at_risk_count:,} student(s) were classified as At Risk.\n\n"
                "Early academic intervention and continuous monitoring "
                "are recommended."
            )
            recommendation_fill = "FEE2E2"
            recommendation_colour = "B91C1C"
        else:
            recommendation_text = (
                "RECOMMENDED ACTION\n\n"
                "No students were classified as At Risk.\n\n"
                "Continue monitoring academic progress."
            )
            recommendation_fill = "DCFCE7"
            recommendation_colour = "166534"

        dashboard_sheet["J10"] = recommendation_text
        dashboard_sheet["J10"].fill = PatternFill(
            "solid",
            fgColor=recommendation_fill,
        )
        dashboard_sheet["J10"].font = Font(
            color=recommendation_colour,
            bold=True,
            size=12,
        )
        dashboard_sheet["J10"].alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )
        dashboard_sheet["J10"].border = Border(
            left=Side(style="medium", color=recommendation_colour),
            right=Side(style="medium", color=recommendation_colour),
            top=Side(style="medium", color=recommendation_colour),
            bottom=Side(style="medium", color=recommendation_colour),
        )

        # Full-width bar chart below all summary content.
        # The image-based chart is consistent across Microsoft Excel and WPS.
        bar_buffer = BytesIO()
        bar_values = [int(counts[category]) for category in category_order]
        bar_colours = ["#22C55E", "#3B82F6", "#F59E0B", "#EF4444"]

        fig, ax = plt.subplots(
            figsize=(11.8, 4.4),
            facecolor="white",
        )
        fig.patch.set_facecolor("white")
        fig.patch.set_edgecolor("#C9D3E6")
        fig.patch.set_linewidth(2.2)
        bars = ax.bar(
            category_order,
            bar_values,
            color=bar_colours,
            width=0.56,
        )
        ax.set_title(
            "Students by Performance Category",
            fontsize=14,
            fontweight="bold",
            pad=14,
        )
        ax.set_ylabel("Number of Students", fontsize=10)
        ax.set_xlabel("Performance Category", fontsize=10)
        for spine in ax.spines.values():
            spine.set_visible(True)
            spine.set_color("#D6DDEB")
            spine.set_linewidth(1.1)
        ax.tick_params(axis="both", labelsize=9)
        ax.grid(False)

        maximum_value = max(bar_values) if bar_values else 1
        ax.set_ylim(0, maximum_value * 1.22)

        for bar, value in zip(bars, bar_values):
            ax.text(
                bar.get_x() + bar.get_width() / 2,
                bar.get_height() + maximum_value * 0.035,
                str(value),
                ha="center",
                va="bottom",
                fontsize=10,
                fontweight="bold",
            )

        fig.patch.set_facecolor("white")
        ax.set_facecolor("white")
        plt.tight_layout(pad=0.8)
        fig.savefig(
            bar_buffer,
            format="png",
            dpi=190,
            bbox_inches="tight",
            facecolor="white",
            edgecolor="#C9D3E6",
        )
        plt.close(fig)
        bar_buffer.seek(0)

        bar_image = XLImage(bar_buffer)
        bar_image.width = 1360
        bar_image.height = 340
        dashboard_sheet.add_image(
            bar_image,
            "A19",
        )

        # Stable row heights.
        row_heights = {
            1: 30,
            2: 18,
            3: 24,
            5: 26,
            6: 26,
            7: 26,
            8: 26,
            9: 10,
            10: 26,
            11: 24,
            12: 24,
            13: 24,
            14: 24,
            15: 24,
            16: 24,
            17: 24,
            18: 10,
            19: 10,
        }
        for row_number, height in row_heights.items():
            dashboard_sheet.row_dimensions[row_number].height = height

        for row_number in range(19, 37):
            dashboard_sheet.row_dimensions[row_number].height = 23

        dashboard_sheet.freeze_panes = "A5"
        dashboard_sheet.page_setup.orientation = "landscape"
        dashboard_sheet.page_setup.fitToWidth = 1
        dashboard_sheet.page_setup.fitToHeight = 1
        dashboard_sheet.sheet_properties.pageSetUpPr.fitToPage = True
        dashboard_sheet.print_area = "A1:O36"
        dashboard_sheet.sheet_view.zoomScale = 85

        # Sheet 4: executive summary.
        summary_sheet = workbook.create_sheet("Summary")
        summary_sheet.sheet_view.showGridLines = False

        summary_sheet.merge_cells("A1:D1")
        summary_sheet["A1"] = "Batch Prediction Summary"
        summary_sheet["A1"].fill = navy_fill
        summary_sheet["A1"].font = Font(
            color="FFFFFF",
            bold=True,
            size=17,
        )
        summary_sheet["A1"].alignment = Alignment(
            horizontal="center",
            vertical="center",
        )
        summary_sheet.row_dimensions[1].height = 34

        summary_sheet.merge_cells("A2:D2")
        summary_sheet["A2"] = (
            f"Total student records processed: {len(export_df):,}"
        )
        summary_sheet["A2"].fill = light_purple_fill
        summary_sheet["A2"].font = Font(
            color="4C1D95",
            bold=True,
        )
        summary_sheet["A2"].alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

        summary_headers = [
            "Performance Category",
            "Number of Students",
            "Percentage",
            "Recommended Action",
        ]
        for column_index, value in enumerate(
            summary_headers,
            start=1,
        ):
            cell = summary_sheet.cell(
                row=4,
                column=column_index,
                value=value,
            )
            cell.fill = purple_fill
            cell.font = white_font
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
            )
            cell.border = Border(
                left=thin_grey,
                right=thin_grey,
                top=thin_grey,
                bottom=thin_grey,
            )

        category_order = [
            "Excellent",
            "Good",
            "Average",
            "At Risk",
        ]
        counts = (
            export_df["Final_Prediction"]
            .value_counts()
            .reindex(category_order, fill_value=0)
        )
        recommendations = {
            "Excellent": "Maintain strong performance",
            "Good": "Continue current progress",
            "Average": "Provide academic guidance",
            "At Risk": "Early intervention required",
        }

        total_students = len(export_df)

        for row_index, category in enumerate(
            category_order,
            start=5,
        ):
            count = int(counts[category])
            percentage = (
                count / total_students
                if total_students
                else 0
            )

            values = [
                category,
                count,
                percentage,
                recommendations[category],
            ]

            for column_index, value in enumerate(
                values,
                start=1,
            ):
                cell = summary_sheet.cell(
                    row=row_index,
                    column=column_index,
                    value=value,
                )
                cell.border = Border(
                    left=thin_grey,
                    right=thin_grey,
                    top=thin_grey,
                    bottom=thin_grey,
                )
                cell.alignment = Alignment(
                    horizontal=(
                        "left"
                        if column_index == 4
                        else "center"
                    ),
                    vertical="center",
                )

            summary_sheet.cell(
                row=row_index,
                column=1,
            ).fill = category_fills[category]
            summary_sheet.cell(
                row=row_index,
                column=1,
            ).font = category_fonts[category]
            summary_sheet.cell(
                row=row_index,
                column=3,
            ).number_format = "0.0%"

        summary_sheet["A10"] = "Priority Focus"
        summary_sheet["A10"].fill = navy_fill
        summary_sheet["A10"].font = white_font
        summary_sheet["B10"] = (
            f"{int(counts['At Risk']):,} At-Risk student(s)"
        )
        summary_sheet["B10"].fill = category_fills["At Risk"]
        summary_sheet["B10"].font = category_fonts["At Risk"]

        summary_sheet.column_dimensions["A"].width = 24
        summary_sheet.column_dimensions["B"].width = 21
        summary_sheet.column_dimensions["C"].width = 16
        summary_sheet.column_dimensions["D"].width = 34
        summary_sheet.freeze_panes = "A4"

        # Open the executive dashboard first.
        workbook.active = workbook.sheetnames.index(
            "Executive Dashboard"
        )

    output.seek(0)
    return output.getvalue()


def make_template_bytes():
    template_file = (
        ROOT
        / "template"
        / "student_batch_prediction_template.xlsx"
    )

    if not template_file.exists():
        raise FileNotFoundError(
            "Batch prediction template was not found in the template folder."
        )

    return template_file.read_bytes()


def validate_batch_data(batch_df):
    required = [
        "Number_of_Subjects",
        "Average_Score",
        "Attendance_Pct",
        "Study_Hours_Per_Day",
        "Previous_CGPA",
    ]
    errors = []
    missing = [column for column in required if column not in batch_df.columns]
    if missing:
        return ["Missing required columns: " + ", ".join(missing)]
    if batch_df.empty:
        return ["The uploaded file does not contain any student records."]

    ranges = {
        "Number_of_Subjects": (1, 12),
        "Average_Score": (0, 100),
        "Attendance_Pct": (0, 100),
        "Study_Hours_Per_Day": (0, 24),
        "Previous_CGPA": (0, 4),
    }

    for column, (minimum, maximum) in ranges.items():
        values = pd.to_numeric(batch_df[column], errors="coerce")
        if values.isna().any():
            rows = (values.isna()).to_numpy().nonzero()[0] + 2
            errors.append(f"{column} contains missing or non-numeric values at Excel row(s): " + ", ".join(map(str, rows[:8])))
            continue
        invalid = ~values.between(minimum, maximum, inclusive="both")
        if invalid.any():
            rows = invalid.to_numpy().nonzero()[0] + 2
            errors.append(f"{column} must be between {minimum} and {maximum}. Invalid Excel row(s): " + ", ".join(map(str, rows[:8])))

    subjects = pd.to_numeric(batch_df["Number_of_Subjects"], errors="coerce")
    non_integer = subjects.notna() & (subjects % 1 != 0)
    if non_integer.any():
        rows = non_integer.to_numpy().nonzero()[0] + 2
        errors.append("Number_of_Subjects must contain whole numbers. Invalid Excel row(s): " + ", ".join(map(str, rows[:8])))
    return errors


def predict_batch(batch_df):
    features = [
        "Average_Score",
        "Attendance_Pct",
        "Study_Hours_Per_Day",
        "Previous_CGPA",
    ]
    result = batch_df.copy()
    best_name = str(evaluation.iloc[0]["Model"])

    for model_name, bundle in models.items():
        model = bundle["model"]
        label_encoder = bundle["label_encoder"]
        encoded = model.predict(batch_df[features]).astype(int)
        labels = label_encoder.inverse_transform(encoded)
        probabilities = model.predict_proba(batch_df[features])
        confidence = probabilities.max(axis=1)
        result[f"{model_name}_Prediction"] = labels
        result[f"{model_name}_Confidence"] = confidence

    result["Final_Prediction"] = result[f"{best_name}_Prediction"]
    result["Final_Confidence"] = result[f"{best_name}_Confidence"]
    result["Best_Model"] = best_name
    return result

def show_cgpa_guide():
    st.markdown(
        """
<div class="cgpa-guide">
<h3 style="margin-top:0;">CGPA Classification Guide</h3>

<div class="cgpa-grid">

<div class="cgpa-card excellent">
<div class="cgpa-title">Excellent</div>
CGPA 3.50 – 4.00
</div>

<div class="cgpa-card good">
<div class="cgpa-title">Good</div>
CGPA 3.00 – 3.49
</div>

<div class="cgpa-card average">
<div class="cgpa-title">Average</div>
CGPA 2.50 – 2.99
</div>

<div class="cgpa-card risk">
<div class="cgpa-title">At Risk</div>
CGPA below 2.50
</div>

</div>
</div>
""",
        unsafe_allow_html=True
    )



df = load_data()
models = load_models()
evaluation = pd.read_csv(RESULTS / "evaluation.csv")

st.sidebar.title("🎓 Student AI")


# Single user navigation (no login / no role restriction)
navigation_options = [
    "🏠 Home",
    "🎯 Prediction",
    "📊 Model Results",
    "🔗 Correlation",
    "📈 Dataset",
    "⭐ Feature Analysis",
    "ℹ️ About"
]


page = st.sidebar.radio(
    "Navigation",
    navigation_options,
    label_visibility="collapsed"
)



page = page.split(" ", 1)[1]

if page != "Prediction":
    st.session_state.pop("prediction_mode", None)

if page != "Prediction":
    st.markdown(
        '<div class="hero">'
        '<h1>Student Performance Prediction</h1>'
        '<p>AI-powered academic performance analysis using KNN, SVM and ANN.</p>'
        '</div>',
        unsafe_allow_html=True
    )

if page == "Home":
    best = evaluation.iloc[0]

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Students", f"{len(df):,}")
    c2.metric("Features", "4")
    c3.metric("Best Model", best["Model"])
    c4.metric("Best Accuracy", f"{best['Accuracy']:.1%}")

    chart = evaluation.melt(
        id_vars="Model",
        value_vars=["Accuracy", "Precision", "Recall", "F1 Score"],
        var_name="Metric",
        value_name="Score"
    )

    fig = px.bar(
        chart,
        x="Model",
        y="Score",
        color="Metric",
        barmode="group",
        range_y=[0, 1],
        title="Model Performance Comparison",
        height=420
    )
    fig.update_layout(
        title=dict(
            x=0.5,
            xanchor="center",
            y=0.96,
            yanchor="top",
        ),
        margin=dict(l=20, r=20, t=75, b=20),
    )

    fig.update_layout(title_x=0.5)
    st.plotly_chart(fig, use_container_width=True)


elif page == "Prediction" and not st.session_state.get("prediction_mode"):
    best_row = evaluation.iloc[0]

    st.markdown('<div class="prediction-hub-shell">', unsafe_allow_html=True)

    st.markdown(
        """
<div class="prediction-welcome">
    <div class="prediction-eyebrow">AI-Powered Academic Decision Support</div>
    <h2>Student Performance Prediction System</h2>
    <p>
        Predict student academic performance using KNN, SVM and ANN.
        Analyse individual student performance and generate AI-based insights.
    </p>
    <p><b>Select a prediction mode below to begin.</b></p>
</div>
""",
        unsafe_allow_html=True,
    )

    stat_col1, stat_col2, stat_col3, stat_col4 = st.columns(4, gap="small")
    stat_col1.metric("🏆 Best Model", str(best_row["Model"]))
    stat_col2.metric("🎯 Best Accuracy", f"{best_row['Accuracy']:.1%}")
    stat_col3.metric("📚 Input Features", "4")
    stat_col4.metric("🤖 ML Models", "3")

    card_col1, card_col2 = st.columns(2, gap="large")

    with card_col1:
        st.markdown(
            """
<div class="prediction-mode-card">
    <div class="prediction-mode-icon">👤</div>
    <div class="prediction-mode-title">Individual Prediction</div>
    <div class="prediction-mode-desc">
        Predict one student's academic performance using KNN, SVM and ANN models.
    </div>
    <div class="prediction-feature-row">
        <div class="prediction-feature-chip">✓ Real-time</div>
        <div class="prediction-feature-chip">✓ 3-model comparison</div>
        <div class="prediction-feature-chip">✓ Excel report</div>
    </div>
</div>
""",
            unsafe_allow_html=True,
        )
        if st.button("🚀 Start Individual Prediction", key="start_individual_prediction",
                     type="primary", use_container_width=True):
            st.session_state["prediction_mode"] = "individual"
            st.rerun()

    with card_col2:
        st.markdown(
            """
<div class="prediction-mode-card">
    <div class="prediction-mode-icon">📁</div>
    <div class="prediction-mode-title">Batch Prediction</div>
    <div class="prediction-mode-desc">
        Upload Excel or CSV files to predict multiple students simultaneously.
        Generate reports and identify at-risk students.
    </div>
    <div class="prediction-feature-row">
        <div class="prediction-feature-chip">✓ Excel / CSV</div>
        <div class="prediction-feature-chip">✓ Dashboard</div>
        <div class="prediction-feature-chip">✓ Risk Detection</div>
    </div>
</div>
""",
            unsafe_allow_html=True,
        )
        if st.button("📂 Start Batch Prediction", key="start_batch_prediction",
                     type="primary", use_container_width=True):
            st.session_state["prediction_mode"] = "batch"
            st.rerun()

    st.markdown(
        """
<div class="why-system">
    <div class="why-system-title">Why Use This System?</div>
    <div class="why-grid">
        <div class="why-item">AI-Powered Prediction</div>
        <div class="why-item">Real-Time Analysis</div>
        <div class="why-item">Batch Processing</div>
        <div class="why-item">Professional Excel Reports</div>
        <div class="why-item">Early At-Risk Identification</div>
    </div>
</div>

<div class="system-footer">
    Student Performance Prediction System • Version 1.0<br>
    Developed using Python, Streamlit and Scikit-learn • BMCS2003 Artificial Intelligence
</div>
</div>
""",
        unsafe_allow_html=True,
    )





def create_batch_template():
    """Create Excel template for batch prediction upload."""
    template_df = pd.DataFrame({
        "Student_ID": ["1234567"],
        "Student_Name": ["Sample Student"],
        "Average_Score": [75.0],
        "Attendance_Pct": [90.0],
        "Study_Hours_Per_Day": [3.0],
        "Previous_CGPA": [3.20],
    })

    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        template_df.to_excel(writer, index=False, sheet_name="Student Data")

    output.seek(0)
    return output.getvalue()


if page == "Prediction" and st.session_state.get("prediction_mode") == "batch":

    st.markdown(
        """
<div class="mode-page-hero">
    <div class="mode-page-eyebrow">Multiple Student Analysis</div>
    <h2>📁 Batch Prediction</h2>
    <p>
        Upload student data in Excel or CSV format to generate predictions for multiple students.
    </p>
</div>
""",
        unsafe_allow_html=True,
    )

    if st.button("← Back to Prediction Modes", key="back_from_batch", type="secondary"):
        st.session_state.pop("prediction_mode", None)
        st.rerun()

    st.download_button(
        "📄 Download Excel Template",
        data=make_template_bytes(),
        file_name="student_batch_prediction_template.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    uploaded_file = st.file_uploader(
        "Upload Excel or CSV file",
        type=["xlsx", "csv"]
    )

    if uploaded_file:
        if uploaded_file.name.endswith(".csv"):
            batch_df = pd.read_csv(uploaded_file)
        else:
            batch_df = pd.read_excel(uploaded_file)

        st.markdown("### 👀 Preview Uploaded Data")
        st.dataframe(
            batch_df.head(100),
            hide_index=True,
            use_container_width=True
        )

        errors = validate_batch_data(batch_df)

        if errors:
            for error in errors:
                st.error(error)
        else:
            if st.button("🚀 Run Batch Prediction", key="run_batch_prediction", use_container_width=True):
                result = predict_batch(batch_df)
                st.session_state["batch_result"] = result
                st.session_state["batch_uploaded_file"] = uploaded_file.name
                st.success("Batch prediction completed successfully!")

    if "batch_result" in st.session_state:
        result = st.session_state["batch_result"]

        # ===== Batch AI Dashboard =====
        st.markdown("### 📊 Batch Prediction Dashboard")

        if "Final_Prediction" in result.columns:
            prediction_col = "Final_Prediction"
        elif "FINAL PREDICTION" in result.columns:
            prediction_col = "FINAL PREDICTION"
        else:
            prediction_col = None

        if prediction_col:
            total = len(result)
            counts = result[prediction_col].value_counts()

            c1, c2, c3, c4 = st.columns(4)
            c1.metric("Total Students", total)
            c2.metric("Excellent", int(counts.get("Excellent", 0)))
            c3.metric("Good", int(counts.get("Good", 0)))
            c4.metric("At Risk", int(counts.get("At Risk", 0)))

            # ===== Large Dashboard Graph =====
            performance_df = counts.reset_index()
            performance_df.columns = ["Performance", "Students"]

            large_fig = px.bar(
                performance_df,
                x="Performance",
                y="Students",
                title="Overall Student Performance Distribution",
                text="Students",
                template="plotly_white",
            )
            large_fig.update_layout(
                height=520,
                xaxis_title="Performance Category",
                yaxis_title="Number of Students",
                title_x=0.5,
            )
            st.plotly_chart(
                large_fig,
                use_container_width=True,
            )

            chart_col1, chart_col2 = st.columns(2)

            with chart_col1:
                pie_df = counts.reset_index()
                pie_df.columns = ["Performance", "Students"]
                fig = px.pie(
                    pie_df,
                    values="Students",
                    names="Performance",
                    title="Prediction Distribution",
                    hole=0.45,
                )
                fig.update_layout(title_x=0.5)
                st.plotly_chart(fig, use_container_width=True)

            with chart_col2:
                bar_df = counts.reset_index()
                bar_df.columns = ["Performance", "Students"]
                fig2 = px.bar(
                    bar_df,
                    x="Performance",
                    y="Students",
                    title="Student Performance Category",
                    text="Students",
                )
                fig2.update_layout(title_x=0.5)
                st.plotly_chart(fig2, use_container_width=True)

            st.markdown("### 📋 Detailed Prediction Results")

        st.dataframe(result, hide_index=True, use_container_width=True)

        excel = make_batch_excel_bytes(result)

        button_col1, button_col2 = st.columns(2, gap="large")

        with button_col1:
            st.download_button(
                "📥 Download Batch Report",
                data=excel,
                file_name="Batch_Prediction_Report.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )

        with button_col2:
            if st.button(
                "🔄 Make Another Prediction",
                key="make_another_batch",
                type="secondary",
                use_container_width=True,
            ):
                st.session_state.pop("batch_result", None)
                st.session_state.pop("batch_uploaded_file", None)
                st.rerun()



elif page == "Prediction" and st.session_state.get("prediction_mode") == "individual":
    st.markdown(
        """
<div class="mode-page-hero">
    <div class="mode-page-eyebrow">Single-Student Analysis</div>
    <h2>👤 Individual Prediction</h2>
    <p>
        Enter one student's academic information to generate real-time
        predictions using KNN, SVM and ANN, then download a formatted Excel result.
    </p>
</div>
""",
        unsafe_allow_html=True,
    )

    if st.button(
        "← Back to Prediction Modes",
        key="back_from_individual",
        type="secondary",
    ):
        st.session_state.pop("prediction_mode", None)
        st.session_state.pop("prediction_result", None)
        st.rerun()

    show_cgpa_guide()

    # Show the latest result at the top of the page.
    # This avoids making the user scroll down after pressing Predict.
    if "prediction_result" in st.session_state:
        saved = st.session_state["prediction_result"]

        st.success(f"Predicted Performance: {saved['prediction']}")
        result_col1, result_col2 = st.columns(2)
        result_col1.metric("Confidence", f"{saved['confidence']:.1%}")
        result_col2.metric("Best Model", saved["best_model"])

        st.dataframe(
            saved["result_df"].assign(
                Confidence=saved["result_df"]["Confidence"].map(
                    lambda x: f"{x:.1%}"
                )
            ),
            hide_index=True,
            use_container_width=True
        )

        button_col1, button_col2 = st.columns(2)

        with button_col1:
            st.download_button(
                "⬇️ Download Result Excel",
                data=saved["report_excel"],
                file_name="individual_prediction_result.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )

        with button_col2:
            if st.button(
                "↻ Make Another Prediction",
                key="make_another_prediction",
                type="secondary",
                use_container_width=True
            ):
                del st.session_state["prediction_result"]
                st.rerun()

    else:
        # All inputs are outside st.form so that changes refresh immediately.
        name = st.text_input("Student Name", key="student_name")

        name_valid = True
        if name.strip():
            name_valid = bool(re.fullmatch(r"[A-Za-z ]+", name.strip()))
            if not name_valid:
                st.error("❌ Student Name can only contain English letters.")

        student_id = st.text_input(
            "Student ID",
            key="student_id",
            placeholder="Enter 7-digit Student ID"
        )

        id_valid = True
        if student_id.strip():
            id_valid = student_id.isdigit() and len(student_id.strip()) == 7
            if not id_valid:
                st.error("❌ Student ID must contain exactly 7 digits.")

        number_of_subjects = st.slider(
            "Number of Subjects",
            min_value=1,
            max_value=12,
            value=5,
            key="number_of_subjects"
        )

        # Subject scores: 0-100 scale
        # Keep .5 values (81.5 stays 81.5), but normal round other decimals
        # Example: 44.4 -> 44, 44.5 -> 44.5, 81.6 -> 82
        def normalize_score(index):
            key = f"subject_{index}"
            value = st.session_state.get(key)

            if value is not None:
                try:
                    value = float(value)
                    decimal = value % 1

                    if abs(decimal - 0.5) < 0.001:
                        converted = value
                    else:
                        converted = round(value)

                    converted = max(0, min(100, converted))
                    st.session_state[key] = converted
                except:
                    st.session_state[key] = None

        scores = []
        columns = st.columns(2)

        for i in range(number_of_subjects):
            with columns[i % 2]:
                key = f"subject_{i}"

                score = st.number_input(
                    f"Subject {i + 1} Score",
                    min_value=0.0,
                    max_value=100.0,
                    value=None,
                    placeholder="Enter score (0-100)",
                    step=0.5,
                    format="%.1f",
                    key=key,
                    on_change=normalize_score,
                    args=(i,)
                )

                scores.append(score)

        valid_scores = [s for s in scores if s is not None]

        if len(valid_scores) == number_of_subjects:
            rounded_scores = [
                s if abs(float(s) % 1 - 0.5) < 0.001 else round(float(s))
                for s in valid_scores
            ]

            raw_average = sum(rounded_scores) / len(rounded_scores)
            average_score = int(math.floor(raw_average + 0.5))
            st.info(f"Calculated Average Score: {average_score}")
        else:
            average_score = None
            st.warning("⚠️ Please complete all subject scores before prediction.")

        attendance = st.slider(
            "Attendance Rate (%)",
            min_value=0.0,
            max_value=100.0,
            value=85.0,
            step=0.5,
            key="attendance"
        )

        study_hours = st.slider(
            "Study Hours Per Day",
            min_value=0.0,
            max_value=12.0,
            value=3.0,
            step=0.1,
            key="study_hours"
        )

        previous_cgpa = st.slider(
            "Previous CGPA",
            min_value=0.0,
            max_value=4.0,
            value=3.0,
            step=0.01,
            key="previous_cgpa"
        )

        submit = st.button(
            "Predict",
            use_container_width=True,
            type="primary"
        )

        if submit:


            if not name.strip():
                st.error("❌ Student Name is required.")
                st.stop()

            if not name_valid:
                st.error("❌ Student Name can only contain English letters.")
                st.stop()

            if not student_id.strip():
                st.error("❌ Student ID is required.")
                st.stop()

            if not id_valid:
                st.error("❌ Student ID must contain exactly 7 digits.")
                st.stop()

            if average_score is None:
                st.error("Please complete all subject scores.")
                st.stop()


            input_df = pd.DataFrame([{
                "Average_Score": average_score,
                "Attendance_Pct": attendance,
                "Study_Hours_Per_Day": study_hours,
                "Previous_CGPA": previous_cgpa,
            }])

            predictions = []
            best_name = evaluation.iloc[0]["Model"]

            for model_name, bundle in models.items():
                model = bundle["model"]
                label_encoder = bundle["label_encoder"]

                pred_code = int(model.predict(input_df)[0])
                pred_label = label_encoder.inverse_transform([pred_code])[0]

                probabilities = model.predict_proba(input_df)[0]
                confidence = float(probabilities[pred_code])

                predictions.append({
                    "Model": model_name,
                    "Prediction": pred_label,
                    "Confidence": confidence,
                })

            result_df = pd.DataFrame(predictions)
            best_row = result_df[result_df["Model"] == best_name].iloc[0]

            prediction_lookup = {
                row["Model"]: row["Prediction"]
                for _, row in result_df.iterrows()
            }

            report = pd.DataFrame([{
                "Student_ID": student_id,
                "Student_Name": name,
                "Number_of_Subjects": number_of_subjects,
                "Average_Score": average_score,
                "Attendance_Pct": attendance,
                "Study_Hours_Per_Day": study_hours,
                "Previous_CGPA": previous_cgpa,
                "KNN_Prediction": prediction_lookup.get("KNN", ""),
                "SVM_Prediction": prediction_lookup.get("SVM", ""),
                "ANN_Prediction": prediction_lookup.get("ANN", ""),
                "Final_Prediction": best_row["Prediction"],
                "Best_Model": best_name,
                "Final_Confidence": best_row["Confidence"],
            }])

            st.session_state["prediction_result"] = {
                "prediction": best_row["Prediction"],
                "confidence": float(best_row["Confidence"]),
                "best_model": best_name,
                "result_df": result_df,
                "report_excel": make_excel_bytes(report),
            }

            st.rerun()

elif page == "Model Results":
    st.subheader("Model Evaluation Dashboard")

    metric_columns = ["Accuracy", "Precision", "Recall", "F1 Score"]
    model_column = "Model"

    best_index = evaluation["Accuracy"].idxmax()
    best_row = evaluation.loc[best_index]
    best_model = str(best_row[model_column])

    st.success(
        f"🏆 Best Model: {best_model} "
        f"with {best_row['Accuracy']:.2%} accuracy"
    )

    st.markdown("### 📊 Model Accuracy Comparison")

    model_colors = {
        "KNN": "#3B82F6",
        "SVM": "#10B981",
        "ANN": "#8B5CF6"
    }

    accuracy_fig = px.bar(
        evaluation,
        x=model_column,
        y="Accuracy",
        color=model_column,
        color_discrete_map=model_colors,
        text=evaluation["Accuracy"].map(lambda value: f"{value:.1%}"),
        title="Accuracy of KNN, SVM and ANN",
        labels={"Accuracy": "Accuracy", model_column: "Model"}
    )
    accuracy_fig.update_traces(textposition="outside")
    accuracy_fig.update_yaxes(
        tickformat=".0%",
        range=[0, min(1.0, float(evaluation["Accuracy"].max()) + 0.12)]
    )
    accuracy_fig.update_layout(
        title=dict(
            x=0.5,
            xanchor="center",
            y=0.96,
            yanchor="top",
        ),
        height=420,
        margin=dict(l=20, r=20, t=75, b=20)
    )
    st.plotly_chart(accuracy_fig, use_container_width=True)

    st.markdown("### 🤖 Individual Model Performance")

    model_order = ["KNN", "SVM", "ANN"]
    model_tabs = st.tabs(model_order)

    for tab, model_name in zip(model_tabs, model_order):
        with tab:
            model_row = evaluation[
                evaluation[model_column].astype(str).str.upper() == model_name
            ]

            if model_row.empty:
                st.warning(f"No evaluation result was found for {model_name}.")
                continue

            model_row = model_row.iloc[0]

            st.markdown(f"#### {model_name} Performance")

            c1, c2, c3, c4 = st.columns(4)

            with c1:
                st.metric("Accuracy", f"{model_row['Accuracy']:.2%}")

            with c2:
                st.metric("Precision", f"{model_row['Precision']:.2%}")

            with c3:
                st.metric("Recall", f"{model_row['Recall']:.2%}")

            with c4:
                st.metric("F1 Score", f"{model_row['F1 Score']:.2%}")

            single_model_data = pd.DataFrame({
                "Metric": metric_columns,
                "Score": [float(model_row[col]) for col in metric_columns]
            })

            model_fig = px.bar(
                single_model_data,
                x="Metric",
                y="Score",
                text=single_model_data["Score"].map(
                    lambda value: f"{value:.1%}"
                ),
                title=f"{model_name} Evaluation Metrics",
                labels={"Score": "Score", "Metric": "Metric"},
                color_discrete_sequence=[model_colors[model_name]]
            )
            model_fig.update_traces(textposition="outside")
            model_fig.update_yaxes(
                tickformat=".0%",
                range=[0, 1.05]
            )
            model_fig.update_layout(
                title=dict(
                    x=0.5,
                    xanchor="center",
                    y=0.96,
                    yanchor="top",
                ),
                height=390,
                margin=dict(l=20, r=20, t=75, b=20)
            )
            st.plotly_chart(model_fig, use_container_width=True)

            if model_name == best_model.upper():
                st.success(
                    f"{model_name} achieved the highest accuracy and was "
                    "selected as the best-performing model."
                )
            else:
                st.info(
                    f"{model_name} was evaluated using Accuracy, Precision, "
                    "Recall and F1 Score."
                )

    st.markdown("### 📋 Complete Model Comparison")

    show = evaluation.copy()
    for col in metric_columns:
        show[col] = show[col].map(lambda value: f"{value:.2%}")

    st.dataframe(
        show,
        hide_index=True,
        use_container_width=True
    )

    st.info(
        "Accuracy shows the overall percentage of correct predictions. "
        "Precision measures how reliable the model's predictions are. "
        "Recall measures how many actual cases were identified, while "
        "F1 Score balances precision and recall."
    )



elif page == "Correlation":
    st.subheader("Feature Correlation & Relevance Analysis")

    st.markdown(
        """
<div class="about-hero">
    <div class="about-hero-title">🔗 Feature Correlation Analysis</div>
    <div class="about-hero-subtitle">
        Explore how available student attributes are associated with academic performance
        and evaluate the relevance of the five features selected for model training.
    </div>
</div>
""",
        unsafe_allow_html=True,
    )

    # ------------------------------------------------------------------
    # Correlation helpers
    # ------------------------------------------------------------------
    def correlation_strength(value):
        absolute_value = abs(float(value))
        if absolute_value < 0.20:
            return "Very weak"
        if absolute_value < 0.40:
            return "Weak"
        if absolute_value < 0.60:
            return "Moderate"
        if absolute_value < 0.80:
            return "Strong"
        return "Very strong"

    def correlation_direction(value):
        value = float(value)
        if value > 0:
            return "Positive"
        if value < 0:
            return "Negative"
        return "No linear"

    # Dataset-level overview.
    numeric_columns = df.select_dtypes(include="number").columns.tolist()

    # IDs are identifiers, not meaningful continuous academic variables.
    overall_numeric_columns = [
        column for column in numeric_columns
        if column.lower() not in {"student_id", "id"}
    ]

    selected_features = [
        "Average_Score",
        "Attendance_Pct",
        "Study_Hours_Per_Day",
        "Previous_CGPA",
    ]
    selected_features = [
        column for column in selected_features
        if column in df.columns
    ]

    overview1, overview2, overview3 = st.columns(3)
    overview1.metric("Students", f"{len(df):,}")
    overview2.metric("Numerical Attributes", len(overall_numeric_columns))
    overview3.metric("Selected Model Features", len(selected_features))

    st.info(
        "This page uses two levels of analysis. Part A examines the available "
        "numerical attributes in the dataset. Part B focuses on the five features "
        "used by the KNN, SVM and ANN prediction models."
    )

    tab_overall, tab_selected = st.tabs(
        [
            "Part A — Overall Dataset Correlation",
            "Part B — Selected Features vs Performance",
        ]
    )

    # ==================================================================
    # PART A: all usable numerical attributes
    # ==================================================================
    with tab_overall:
        st.markdown("### Part A — Overall Dataset Correlation")
        st.caption(
            "Pearson correlation is used here to examine linear relationships "
            "between the available numerical attributes. Student_ID is excluded "
            "because it is an identifier rather than an academic measurement."
        )

        if len(overall_numeric_columns) < 2:
            st.warning(
                "At least two numerical attributes are required to calculate "
                "the overall correlation matrix."
            )
        else:
            overall_corr = (
                df[overall_numeric_columns]
                .apply(pd.to_numeric, errors="coerce")
                .corr(method="pearson")
                .round(2)
            )

            overall_fig = px.imshow(
                overall_corr,
                text_auto=".2f",
                aspect="auto",
                color_continuous_scale="RdBu_r",
                zmin=-1,
                zmax=1,
                title="Overall Numerical Attribute Correlation Matrix",
            )
            overall_fig.update_layout(
                title=dict(
                    x=0.5,
                    xanchor="center",
                    y=0.98,
                    yanchor="top",
                ),
                height=max(560, 55 * len(overall_corr.columns)),
                margin=dict(l=30, r=30, t=85, b=30),
                coloraxis_colorbar=dict(title="Pearson r"),
            )
            overall_fig.update_xaxes(side="bottom")
            st.plotly_chart(overall_fig, use_container_width=True)

            with st.expander("View Overall Correlation Values"):
                st.dataframe(
                    overall_corr,
                    hide_index=True,
                    use_container_width=True,
                )

            st.markdown("#### Numerical Attributes Included")
            st.write(", ".join(overall_numeric_columns))

            if "Final_CGPA" in overall_corr.columns:
                final_cgpa_corr = (
                    overall_corr["Final_CGPA"]
                    .drop(labels=["Final_CGPA"], errors="ignore")
                    .dropna()
                    .sort_values(key=lambda series: series.abs(), ascending=False)
                )

                if not final_cgpa_corr.empty:
                    overall_rank_df = final_cgpa_corr.rename(
                        "Correlation"
                    ).reset_index()
                    overall_rank_df.columns = ["Attribute", "Correlation"]

                    overall_rank_fig = px.bar(
                        overall_rank_df,
                        x="Correlation",
                        y="Attribute",
                        orientation="h",
                        text=overall_rank_df["Correlation"].map(
                            lambda value: f"{value:+.2f}"
                        ),
                        title="Numerical Attributes Associated with Final CGPA",
                        range_x=[-1, 1],
                    )
                    overall_rank_fig.update_traces(textposition="outside")
                    overall_rank_fig.update_layout(
                        title=dict(x=0.5, xanchor="center"),
                        yaxis=dict(
                            categoryorder="array",
                            categoryarray=overall_rank_df["Attribute"].tolist()[::-1],
                        ),
                        height=max(400, 48 * len(overall_rank_df)),
                        margin=dict(l=20, r=30, t=75, b=20),
                    )
                    st.plotly_chart(overall_rank_fig, use_container_width=True)

    # ==================================================================
    # PART B: selected model features vs ordinal performance category
    # ==================================================================
    with tab_selected:
        st.markdown("### Part B — Selected Features vs Performance")
        st.caption(
            "Performance Category is ordinally encoded only for this analysis: "
            "At Risk = 0, Average = 1, Good = 2, Excellent = 3. "
            "Spearman correlation is used because the target categories have "
            "a natural order."
        )

        performance_map = {
            "At Risk": 0,
            "Average": 1,
            "Good": 2,
            "Excellent": 3,
        }

        analysis_df = df[selected_features + ["Performance_Category"]].copy()
        analysis_df["Performance_Score"] = (
            analysis_df["Performance_Category"]
            .astype(str)
            .map(performance_map)
        )

        for feature in selected_features:
            analysis_df[feature] = pd.to_numeric(
                analysis_df[feature],
                errors="coerce",
            )

        focused_columns = selected_features + ["Performance_Score"]
        focused_corr = (
            analysis_df[focused_columns]
            .corr(method="spearman")
            .round(2)
        )

        # --------------------------------------------------------------
        # Interactive relationship chart based on the actual student dataset
        # --------------------------------------------------------------
        st.markdown("### Student Dataset Correlation Pattern")
        st.caption(
            "Select one model feature to examine its relationship with student "
            "performance. The chart uses the actual student records. Performance "
            "is displayed on an ordinal scale (At Risk → Average → Good → Excellent)."
        )

        feature_labels = {
            "Number_of_Subjects": "Number of Subjects",
            "Average_Score": "Average Score",
            "Attendance_Pct": "Attendance Percentage",
            "Study_Hours_Per_Day": "Study Hours per Day",
            "Previous_CGPA": "Previous CGPA",
        }

        available_feature_options = [
            feature for feature in [
                "Previous_CGPA",
                "Average_Score",
                "Attendance_Pct",
                "Study_Hours_Per_Day",
                "Number_of_Subjects",
            ]
            if feature in selected_features
        ]

        selected_feature = st.selectbox(
            "Select Feature to Analyse",
            options=available_feature_options,
            format_func=lambda feature: feature_labels.get(feature, feature),
            key="correlation_feature_selector",
        )

        feature_data = analysis_df[
            [selected_feature, "Performance_Score", "Performance_Category"]
        ].dropna().copy()

        # Display-only jitter reduces overlap between the four ordinal categories.
        jitter_sequence = [
            ((position % 13) - 6) * 0.022
            for position in range(len(feature_data))
        ]
        feature_data["Performance_Display"] = (
            feature_data["Performance_Score"].astype(float)
            + pd.Series(jitter_sequence, index=feature_data.index)
        )

        rho = float(
            analysis_df[[selected_feature, "Performance_Score"]]
            .corr(method="spearman")
            .iloc[0, 1]
        )
        strength = correlation_strength(rho)
        direction = correlation_direction(rho)

        summary_col1, summary_col2, summary_col3 = st.columns(3)
        summary_col1.metric(
            "Selected Feature",
            feature_labels.get(selected_feature, selected_feature),
        )
        summary_col2.metric(
            "Spearman Correlation",
            f"{rho:+.2f}",
        )
        summary_col3.metric(
            "Relationship",
            f"{strength} {direction}",
        )

        relationship_fig = px.scatter(
            feature_data,
            x=selected_feature,
            y="Performance_Display",
            color="Performance_Category",
            category_orders={
                "Performance_Category": [
                    "At Risk",
                    "Average",
                    "Good",
                    "Excellent",
                ]
            },
            title=(
                f"{feature_labels.get(selected_feature, selected_feature)} "
                f"vs Student Performance"
                f"<br><sup>ρ = {rho:+.2f} • {strength} {direction.lower()} correlation</sup>"
            ),
            labels={
                selected_feature: feature_labels.get(
                    selected_feature,
                    selected_feature,
                ),
                "Performance_Display": "Performance Category",
                "Performance_Category": "Category",
            },
            opacity=0.42,
        )

        relationship_fig.update_traces(
            marker=dict(size=6),
            selector=dict(mode="markers"),
        )

        # Add a simple visual trend line without additional dependencies.
        x_values = pd.to_numeric(
            feature_data[selected_feature],
            errors="coerce",
        )
        y_values = pd.to_numeric(
            feature_data["Performance_Score"],
            errors="coerce",
        )
        valid_mask = x_values.notna() & y_values.notna()
        x_values = x_values[valid_mask]
        y_values = y_values[valid_mask]

        if (
            len(x_values) > 1
            and float(((x_values - x_values.mean()) ** 2).sum()) > 0
        ):
            x_mean = float(x_values.mean())
            y_mean = float(y_values.mean())
            denominator = float(((x_values - x_mean) ** 2).sum())
            slope = float(
                ((x_values - x_mean) * (y_values - y_mean)).sum()
                / denominator
            )
            intercept = y_mean - slope * x_mean

            line_x = [
                float(x_values.min()),
                float(x_values.max()),
            ]
            line_y = [
                slope * line_x[0] + intercept,
                slope * line_x[1] + intercept,
            ]

            relationship_fig.add_scatter(
                x=line_x,
                y=line_y,
                mode="lines",
                name="Trend",
                line=dict(width=4),
            )

        relationship_fig.update_yaxes(
            tickmode="array",
            tickvals=[0, 1, 2, 3],
            ticktext=[
                "At Risk",
                "Average",
                "Good",
                "Excellent",
            ],
            range=[-0.35, 3.35],
        )

        relationship_fig.update_layout(
            title=dict(
                x=0.5,
                xanchor="center",
                y=0.97,
                yanchor="top",
            ),
            height=560,
            margin=dict(l=30, r=30, t=95, b=35),
            legend=dict(
                orientation="h",
                yanchor="bottom",
                y=-0.22,
                xanchor="center",
                x=0.5,
            ),
        )

        st.plotly_chart(
            relationship_fig,
            use_container_width=True,
            key="selected_feature_relationship_chart",
        )

        # Feature-specific interpretation.
        if abs(rho) < 0.05:
            interpretation_text = (
                f"{feature_labels.get(selected_feature, selected_feature)} "
                "shows no meaningful monotonic relationship with student "
                "performance in this dataset."
            )
        else:
            interpretation_text = (
                f"{feature_labels.get(selected_feature, selected_feature)} "
                f"shows a {strength.lower()} {direction.lower()} association "
                f"with student performance (ρ = {rho:+.2f})."
            )

        st.info(interpretation_text)

        st.markdown(
            """
<div class="about-card about-purple">
    <h4>How to Read This Chart</h4>
    <div class="about-list-item">
        An upward trend indicates a positive association with student performance,
        while a downward trend indicates a negative association.
    </div>
    <div class="about-list-item">
        A flatter pattern indicates a weaker individual relationship. The Spearman
        coefficient shown above the chart is calculated from the actual dataset.
    </div>
</div>
""",
            unsafe_allow_html=True,
        )

        st.markdown("### Selected Features Correlation Matrix")

        focused_fig = px.imshow(
            focused_corr,
            text_auto=".2f",
            aspect="auto",
            color_continuous_scale="RdBu_r",
            zmin=-1,
            zmax=1,
            title="Selected Features and Performance Correlation Matrix",
        )
        focused_fig.update_layout(
            title=dict(
                x=0.5,
                xanchor="center",
                y=0.98,
                yanchor="top",
            ),
            height=600,
            margin=dict(l=30, r=30, t=85, b=30),
            coloraxis_colorbar=dict(title="Spearman ρ"),
        )
        st.plotly_chart(focused_fig, use_container_width=True)

        performance_corr = (
            focused_corr["Performance_Score"]
            .drop(labels=["Performance_Score"], errors="ignore")
            .dropna()
            .sort_values(key=lambda series: series.abs(), ascending=False)
        )

        ranking_df = performance_corr.rename(
            "Correlation"
        ).reset_index()
        ranking_df.columns = ["Feature", "Correlation"]
        ranking_df["Strength"] = ranking_df["Correlation"].map(
            correlation_strength
        )
        ranking_df["Direction"] = ranking_df["Correlation"].map(
            correlation_direction
        )

        st.markdown("### Feature–Performance Correlation Ranking")

        ranking_fig = px.bar(
            ranking_df,
            x="Correlation",
            y="Feature",
            orientation="h",
            text=ranking_df["Correlation"].map(
                lambda value: f"{value:+.2f}"
            ),
            title="Correlation of Selected Features with Student Performance",
            range_x=[-1, 1],
        )
        ranking_fig.update_traces(textposition="outside")
        ranking_fig.update_layout(
            title=dict(
                x=0.5,
                xanchor="center",
                y=0.96,
                yanchor="top",
            ),
            yaxis=dict(
                categoryorder="array",
                categoryarray=ranking_df["Feature"].tolist()[::-1],
            ),
            height=430,
            margin=dict(l=20, r=30, t=75, b=20),
        )
        st.plotly_chart(ranking_fig, use_container_width=True)

        if not ranking_df.empty:
            strongest = ranking_df.iloc[0]
            weakest = ranking_df.iloc[-1]

            metric1, metric2, metric3 = st.columns(3)
            metric1.metric(
                "Strongest Association",
                strongest["Feature"].replace("_", " "),
            )
            metric2.metric(
                "Strongest Correlation",
                f"{strongest['Correlation']:+.2f}",
            )
            metric3.metric(
                "Relationship Strength",
                strongest["Strength"],
            )

            st.markdown("### Detailed Interpretation")
            interpretation_table = ranking_df.copy()
            interpretation_table["Feature"] = (
                interpretation_table["Feature"]
                .str.replace("_", " ", regex=False)
            )
            interpretation_table["Correlation"] = (
                interpretation_table["Correlation"]
                .map(lambda value: f"{value:+.2f}")
            )

            st.dataframe(
                interpretation_table,
                hide_index=True,
                use_container_width=True,
            )

            strongest_direction = strongest["Direction"].lower()
            weakest_direction = weakest["Direction"].lower()

            st.success(
                f"The strongest association with student performance is "
                f"{strongest['Feature'].replace('_', ' ')} "
                f"(ρ = {strongest['Correlation']:+.2f}), representing a "
                f"{strongest['Strength'].lower()} {strongest_direction} relationship."
            )

            st.info(
                f"The weakest association among the selected model features is "
                f"{weakest['Feature'].replace('_', ' ')} "
                f"(ρ = {weakest['Correlation']:+.2f}), representing a "
                f"{weakest['Strength'].lower()} {weakest_direction} relationship."
            )

            st.markdown(
                f"""
<div class="about-card about-blue">
    <h4>Key Finding</h4>
    <div class="about-list-item">
        <b>{strongest['Feature'].replace('_', ' ')}</b> shows the strongest statistical
        association with student performance (ρ = {strongest['Correlation']:+.2f}).
    </div>
    <div class="about-list-item">
        <b>{weakest['Feature'].replace('_', ' ')}</b> shows the weakest individual
        association (ρ = {weakest['Correlation']:+.2f}).
    </div>
    <div class="about-list-item">
        The selected features contribute different levels of association and should be
        interpreted together rather than individually. A weak individual correlation
        does not automatically mean a feature has no predictive value when combined
        with other variables in a machine learning model.
    </div>
</div>
""",
                unsafe_allow_html=True,
            )

        st.markdown("### How to Interpret the Correlation")
        interpretation_guide = pd.DataFrame(
            {
                "|Correlation|": [
                    "0.00 – 0.19",
                    "0.20 – 0.39",
                    "0.40 – 0.59",
                    "0.60 – 0.79",
                    "0.80 – 1.00",
                ],
                "Interpretation": [
                    "Very weak",
                    "Weak",
                    "Moderate",
                    "Strong",
                    "Very strong",
                ],
            }
        )
        st.dataframe(
            interpretation_guide,
            hide_index=True,
            use_container_width=True,
        )

        st.warning(
            "Correlation measures statistical association, not causation. "
            "A higher absolute correlation coefficient indicates a stronger "
            "association with performance, but it does not prove that the "
            "feature directly causes the student's performance."
        )


elif page == "Dataset":
    st.subheader("Dataset Analysis")

    c1, c2 = st.columns(2)

    with c1:
        fig = px.histogram(
            df,
            x="Average_Score",
            nbins=25,
            title="Average Score Distribution"
        )
        fig.update_layout(
            title=dict(
                x=0.5,
                xanchor="center",
                y=0.96,
                yanchor="top",
            ),
            margin=dict(l=20, r=20, t=75, b=20),
        )
        fig.update_layout(title_x=0.5)
    st.plotly_chart(fig, use_container_width=True)

    with c2:
        fig = px.scatter(
            df,
            x="Attendance_Pct",
            y="Average_Score",
            color="Performance_Category",
            title="Attendance vs Average Score"
        )
        fig.update_layout(
            title=dict(
                x=0.5,
                xanchor="center",
                y=0.96,
                yanchor="top",
            ),
            margin=dict(l=20, r=20, t=75, b=20),
        )
        fig.update_layout(title_x=0.5)
    st.plotly_chart(fig, use_container_width=True)

    st.markdown("### Complete Student Dataset")

    dataset_info_col, dataset_download_col = st.columns([1, 2])

    with dataset_info_col:
        st.metric(
            "Total Students",
            f"{len(df):,}",
        )

    with dataset_download_col:
        st.caption(
            "Browse all student records below or download the complete dataset "
            "as a CSV file."
        )
        st.download_button(
            label="📥 Download Complete Dataset",
            data=df.to_csv(index=False).encode("utf-8"),
            file_name="student_performance_dataset.csv",
            mime="text/csv",
            use_container_width=True,
        )

    st.dataframe(
        df,
        hide_index=True,
        use_container_width=True,
        height=700,
    )



elif page == "Feature Analysis":

    st.markdown(
        """
<div class="mode-page-hero">
    <div class="mode-page-eyebrow">
        Feature Engineering
    </div>
    <h2>⭐ Feature Analysis</h2>
    <p>
    Understand the key factors selected for student performance prediction.
    </p>
</div>
""",
        unsafe_allow_html=True
    )

    st.subheader("🔍 Selected Model Features")

    st.markdown(
        """
<div class="about-card about-blue">

<h3>📚 Academic Performance</h3>

<div class="about-list-item">
<b>Average Score</b><br>
Overall academic achievement indicator based on student assessment results.
</div>

<div class="about-list-item">
<b>Previous CGPA</b><br>
Represents previous academic performance and learning consistency.
</div>

<h3>📈 Learning Behaviour</h3>

<div class="about-list-item">
<b>Attendance Rate</b><br>
Shows student participation and class engagement level.
</div>

<div class="about-list-item">
<b>Study Hours Per Day</b><br>
Represents student's daily learning effort and study habits.
</div>

</div>
""",
        unsafe_allow_html=True
    )

    st.subheader("⭐ Feature Importance Overview")

    importance_data = pd.DataFrame({
        "Feature": [
            "Previous CGPA",
            "Average Score",
            "Study Hours Per Day",
            "Attendance Rate"
        ],
        "Importance": [
            0.69,
            0.55,
            0.33,
            0.36
        ]
    })

    fig = px.bar(
        importance_data,
        x="Importance",
        y="Feature",
        orientation="h",
        title="Relative Feature Contribution"
    )

    fig.update_layout(
        title=dict(
            x=0.5,
            xanchor="center"
        ),
        xaxis_title="Importance Score",
        yaxis_title=""
    )

    fig.update_layout(title_x=0.5)
    st.plotly_chart(fig, use_container_width=True)

elif page == "About":
    st.subheader("About This Assignment")

    st.markdown(
        """
<div class="about-hero">
    <div class="about-hero-title">🎓 Student Performance Prediction</div>
    <div class="about-hero-subtitle">
        A supervised machine learning classification assignment for predicting overall student academic performance.
    </div>
</div>
""",
        unsafe_allow_html=True
    )

    st.markdown("### 📋 Assignment Overview")

    overview_col1, overview_col2, overview_col3, overview_col4 = st.columns(4)

    with overview_col1:
        st.metric("Assignment Type", "Supervised")

    with overview_col2:
        st.metric("Task", "Classification")

    with overview_col3:
        st.metric("Students", f"{len(df):,}")

    with overview_col4:
        st.metric("Features", "4")

    st.markdown("### 📊 Input Features and Models")

    left_col, right_col = st.columns(2)

    with left_col:
        st.markdown(
            """
<div class="about-card about-blue">
    <h4>📋 Input Features</h4>
    <div class="about-list-item">1. Average Score</div>
    <div class="about-list-item">2. Attendance Rate</div>
    <div class="about-list-item">3. Study Hours Per Day</div>
    <div class="about-list-item">4. Previous CGPA</div>
</div>
""",
            unsafe_allow_html=True
        )

    with right_col:
        st.markdown(
            """
<div class="about-card about-green">
    <h4>🤖 Machine Learning Models</h4>
    <div class="about-model"><b>KNN</b> — K-Nearest Neighbours</div>
    <div class="about-model"><b>SVM</b> — Support Vector Machine</div>
    <div class="about-model"><b>ANN</b> — Artificial Neural Network</div>
</div>
""",
            unsafe_allow_html=True
        )

    st.markdown("### 🎯 Target Classes")

    target_col1, target_col2, target_col3, target_col4 = st.columns(4)

    with target_col1:
        st.markdown(
            """
<div class="about-target about-green">
    <div class="about-target-title">Excellent</div>
    <div class="about-target-range">CGPA 3.50 – 4.00</div>
</div>
""",
            unsafe_allow_html=True
        )

    with target_col2:
        st.markdown(
            """
<div class="about-target about-blue">
    <div class="about-target-title">Good</div>
    <div class="about-target-range">CGPA 3.00 – 3.49</div>
</div>
""",
            unsafe_allow_html=True
        )

    with target_col3:
        st.markdown(
            """
<div class="about-target about-yellow">
    <div class="about-target-title">Average</div>
    <div class="about-target-range">CGPA 2.50 – 2.99</div>
</div>
""",
            unsafe_allow_html=True
        )

    with target_col4:
        st.markdown(
            """
<div class="about-target about-red">
    <div class="about-target-title">At Risk</div>
    <div class="about-target-range">CGPA below 2.50</div>
</div>
""",
            unsafe_allow_html=True
        )

    st.markdown("### 📚 Dataset Information")

    info_col1, info_col2 = st.columns(2)

    with info_col1:
        st.markdown(
            """
<div class="about-card about-purple">
    <h4>🗂️ Dataset Source</h4>
    <div class="about-list-item">
        University Student Performance & Habits Dataset from Kaggle
    </div>
    <h4 style="margin-top:14px;">🧩 Feature Engineering</h4>
    <div class="about-list-item">Number of Subjects</div>
    <div class="about-list-item">Average Score</div>
</div>
""",
            unsafe_allow_html=True
        )

    with info_col2:
        st.markdown(
            """
<div class="about-card about-orange">
    <h4>⚠️ Important Note</h4>
    <div class="about-list-item">
        Final CGPA is used only during model training to generate the target class.
    </div>
    <div class="about-list-item">
        It is not included as a prediction input because that would cause data leakage.
    </div>
</div>
""",
            unsafe_allow_html=True
        )



st.sidebar.markdown(
    """
    <div class="sidebar-footer">
        🎓 Student AI • Version 1.0 • RIS Group 5 • ©2026
    </div>
    """,
    unsafe_allow_html=True
)

# FINAL ULTIMATE VERSION FEATURES
# - Prediction page supports single and batch prediction
# - Number_of_Subjects retained as profile information
# - Final ML features:
#   Average_Score
#   Attendance_Pct
#   Study_Hours_Per_Day
#   Previous_CGPA
# - Feature analysis and model comparison included
# - Export prediction reports supported


# FINAL SUBMISSION VERSION
# Design goals:
# - Original UI style preserved
# - Single student prediction + batch workflow
# - KNN / SVM / ANN comparison
# - Feature selection explanation
# - Number_of_Subjects retained as profile information
# - Final ML features:
#   Average_Score
#   Attendance_Pct
#   Study_Hours_Per_Day
#   Previous_CGPA
